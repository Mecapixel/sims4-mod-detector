#!/usr/bin/env python3
"""
Sims 4 Mod Conflict Detector v8.0
Complete rewrite with proper DBPF package parsing, real TGI-based conflict
detection, actual LOD/GEOM analysis, and reliable broken CC identification.
Now auto-generates a formatted Excel cleanup checklist after every scan.
NEW: Scarlet Realm Mod Checker CSV cross-reference — drag & drop your daily
mod checker list to flag broken, outdated, and updated mods in your collection.

Requires: Python 3.10+
Optional: tqdm (pip install tqdm) for progress bars
Optional: openpyxl (pip install openpyxl) for Excel cleanup checklist
"""
import os
import sys
import struct
import zlib
import zipfile
import io
import argparse
import hashlib
import re
import json
import csv
import logging
from pathlib import Path
from collections import defaultdict
from datetime import datetime
from multiprocessing import Pool, cpu_count, freeze_support
from typing import Optional

# Version
VERSION = "8.0"
BUILD_DATE = "2026-02-22"

# Windows UTF-8 fix
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

# Optional tqdm import with fallback
try:
    from tqdm import tqdm
except ImportError:
    def tqdm(iterable, **kwargs):
        """Fallback: plain iterator with periodic status prints."""
        desc = kwargs.get("desc", "Processing")
        total = kwargs.get("total", None)
        if total is None:
            try:
                total = len(iterable)
            except TypeError:
                total = None
        count = 0
        step = max(1, (total // 20)) if total else 500
        for item in iterable:
            count += 1
            if count % step == 0 or count == total:
                if total:
                    print(f"\r  {desc}: {count}/{total} ({100*count//total}%)", end="", flush=True)
                else:
                    print(f"\r  {desc}: {count}...", end="", flush=True)
            yield item
        if count > 0:
            print()


# ═══════════════════════════════════════════════════════════════════════════════
# DBPF PARSER — Proper binary parsing of Sims 4 .package files
# ═══════════════════════════════════════════════════════════════════════════════

class DBPFError(Exception):
    """Raised when a .package file cannot be parsed."""
    pass


class DBPFEntry:
    """A single resource entry from a DBPF index table."""
    __slots__ = ("res_type", "group", "instance", "offset", "file_size",
                 "mem_size", "compressed", "compression_type")

    def __init__(self, res_type, group, instance, offset, file_size,
                 mem_size, compressed, compression_type):
        self.res_type = res_type
        self.group = group
        self.instance = instance
        self.offset = offset
        self.file_size = file_size
        self.mem_size = mem_size
        self.compressed = compressed
        self.compression_type = compression_type

    @property
    def tgi(self):
        return (self.res_type, self.group, self.instance)

    @property
    def type_hex(self):
        return f"0x{self.res_type:08X}"

    @property
    def instance_hex(self):
        return f"0x{self.instance:016X}"

    def __repr__(self):
        return (f"DBPFEntry(type={self.type_hex}, group=0x{self.group:08X}, "
                f"inst={self.instance_hex}, size={self.file_size})")


class DBPFParser:
    """
    Parses Sims 4 DBPF 2.0/2.1 .package files.

    Reads the binary header and index table to extract resource entries with
    their Type/Group/Instance identifiers, offsets, and sizes. Can decompress
    individual resources (ZLIB) on demand.
    """

    # ── Well-known Sims 4 resource type IDs ──
    TUNING_XML    = 0x62ECC59A   # XML tuning definitions
    SIMDATA       = 0x545AC67A   # Compiled SimData
    CAS_PART      = 0x034AEECB   # CAS part definition
    GEOM          = 0x015A1849   # Geometry / mesh
    CLIP          = 0x6B20C4F3   # Animation clip
    OBJDEF        = 0x319E4F1D   # Object definition
    IMG_DDS       = 0x00B2D882   # DDS image
    IMG_PNG       = 0x2F7D0004   # PNG image
    THUMB_SM      = 0x3C1AF1F2   # Small thumbnail
    THUMB_MD      = 0xCD9DE247   # Medium thumbnail
    THUMB_LG      = 0x5B282D45   # Large thumbnail
    STBL          = 0x220557DA   # String table
    MODL          = 0x01661233   # Model / mesh reference
    LITE          = 0x03B4C61D   # Lighting data
    JAZZ          = 0x02D5DF13   # Jazz animation state machine
    CASP          = 0x034AEECB   # CAS Part (alias)
    REGION_MAP    = 0xAC16FBEC   # Region map
    WORLD_OBJ     = 0xD3044521   # World object
    SLOT          = 0xD3044521   # Slot data

    # Compression types
    COMPRESS_NONE    = 0x0000
    COMPRESS_ZLIB    = 0x5A42   # "ZB"
    COMPRESS_INTERNAL = 0xFFFF
    COMPRESS_DELETED  = 0xFFE0

    # Friendly names for common types
    TYPE_NAMES = {
        0x62ECC59A: "Tuning XML",
        0x545AC67A: "SimData",
        0x034AEECB: "CAS Part",
        0x015A1849: "GEOM (Mesh)",
        0x6B20C4F3: "Animation Clip",
        0x319E4F1D: "Object Definition",
        0x00B2D882: "DDS Image",
        0x2F7D0004: "PNG Image",
        0x3C1AF1F2: "Thumbnail (S)",
        0xCD9DE247: "Thumbnail (M)",
        0x5B282D45: "Thumbnail (L)",
        0x220557DA: "String Table",
        0x01661233: "Model",
        0x02D5DF13: "Jazz (ASM)",
    }

    HEADER_SIZE = 96

    def __init__(self, filepath):
        self.filepath = Path(filepath)
        self.entries: list[DBPFEntry] = []
        self.entry_count = 0
        self.index_offset = 0
        self.index_size = 0
        self.major_version = 0
        self.minor_version = 0
        self._parsed = False

    def parse(self) -> bool:
        """Parse the DBPF header and index table. Returns True on success."""
        try:
            with open(self.filepath, "rb") as f:
                header = f.read(self.HEADER_SIZE)

            if len(header) < self.HEADER_SIZE:
                raise DBPFError("File too small for DBPF header")

            magic = header[0:4]
            if magic != b"DBPF":
                raise DBPFError(f"Bad magic: {magic!r} (expected b'DBPF')")

            self.major_version = struct.unpack_from("<I", header, 4)[0]
            self.minor_version = struct.unpack_from("<I", header, 8)[0]

            if self.major_version != 2:
                raise DBPFError(f"Unsupported DBPF version {self.major_version}.{self.minor_version}")

            self.entry_count = struct.unpack_from("<I", header, 0x24)[0]
            self.index_size = struct.unpack_from("<I", header, 0x2C)[0]
            self.index_offset = struct.unpack_from("<I", header, 0x40)[0]

            if self.entry_count == 0:
                self._parsed = True
                return True

            # Sanity checks
            file_size = self.filepath.stat().st_size
            if self.index_offset >= file_size:
                raise DBPFError(f"Index offset {self.index_offset} beyond file size {file_size}")
            if self.index_offset + self.index_size > file_size + 1024:
                raise DBPFError("Index extends beyond file")

            # Read and parse index table
            with open(self.filepath, "rb") as f:
                f.seek(self.index_offset)
                index_data = f.read(self.index_size)

            self._parse_index(index_data)
            self._parsed = True
            return True

        except DBPFError:
            raise
        except Exception as e:
            raise DBPFError(f"Parse error: {e}") from e

    def _parse_index(self, data: bytes):
        """Parse the index table with flag-based constant fields."""
        if len(data) < 4:
            raise DBPFError("Index data too short")

        pos = 0
        flags = struct.unpack_from("<I", data, pos)[0]
        pos += 4

        # Read constant TGI fields (stored once if flagged)
        const_type = const_group = const_inst_hi = const_inst_lo = 0
        if flags & 0x01:
            const_type = struct.unpack_from("<I", data, pos)[0]
            pos += 4
        if flags & 0x02:
            const_group = struct.unpack_from("<I", data, pos)[0]
            pos += 4
        if flags & 0x04:
            const_inst_hi = struct.unpack_from("<I", data, pos)[0]
            pos += 4
        if flags & 0x08:
            const_inst_lo = struct.unpack_from("<I", data, pos)[0]
            pos += 4

        entries_data_start = pos
        remaining = len(data) - entries_data_start
        if self.entry_count > 0 and remaining > 0:
            per_entry = remaining // self.entry_count
        else:
            per_entry = 0

        for _ in range(self.entry_count):
            if pos + 12 > len(data):  # Minimum: offset + size + memsize
                break

            try:
                # Type
                if flags & 0x01:
                    res_type = const_type
                else:
                    res_type = struct.unpack_from("<I", data, pos)[0]
                    pos += 4

                # Group
                if flags & 0x02:
                    group = const_group
                else:
                    group = struct.unpack_from("<I", data, pos)[0]
                    pos += 4

                # Instance High
                if flags & 0x04:
                    inst_hi = const_inst_hi
                else:
                    inst_hi = struct.unpack_from("<I", data, pos)[0]
                    pos += 4

                # Instance Low
                if flags & 0x08:
                    inst_lo = const_inst_lo
                else:
                    inst_lo = struct.unpack_from("<I", data, pos)[0]
                    pos += 4

                instance = (inst_hi << 32) | inst_lo

                # Position, size, memory size
                offset = struct.unpack_from("<I", data, pos)[0]
                pos += 4

                size_raw = struct.unpack_from("<I", data, pos)[0]
                pos += 4

                compressed = bool(size_raw & 0x80000000)
                file_size = size_raw & 0x7FFFFFFF

                mem_size = struct.unpack_from("<I", data, pos)[0]
                pos += 4

                compression_type = self.COMPRESS_NONE
                if compressed and pos + 4 <= len(data) and per_entry > 16:
                    # Some formats store extra compression info
                    comp_raw = struct.unpack_from("<I", data, pos)[0]
                    compression_type = comp_raw >> 16
                    pos += 4

                self.entries.append(DBPFEntry(
                    res_type=res_type,
                    group=group,
                    instance=instance,
                    offset=offset,
                    file_size=file_size,
                    mem_size=mem_size,
                    compressed=compressed,
                    compression_type=compression_type,
                ))
            except struct.error:
                break  # Ran out of data

    def read_resource(self, entry: DBPFEntry) -> Optional[bytes]:
        """Read and decompress a single resource from the package."""
        try:
            with open(self.filepath, "rb") as f:
                f.seek(entry.offset)
                raw = f.read(entry.file_size)

            if not entry.compressed:
                return raw

            # Try ZLIB decompression
            if entry.compression_type in (self.COMPRESS_ZLIB, self.COMPRESS_NONE, 0):
                try:
                    return zlib.decompress(raw)
                except zlib.error:
                    pass

            # Try raw deflate (no header)
            try:
                return zlib.decompress(raw, -zlib.MAX_WBITS)
            except zlib.error:
                pass

            # Try with zlib header
            try:
                return zlib.decompress(raw, zlib.MAX_WBITS)
            except zlib.error:
                pass

            # Return raw if decompression fails
            return raw

        except Exception:
            return None

    def get_by_type(self, type_id: int) -> list[DBPFEntry]:
        """Return all entries matching a resource type."""
        return [e for e in self.entries if e.res_type == type_id]

    def get_resource_type_counts(self) -> dict[int, int]:
        """Count entries per resource type."""
        counts = defaultdict(int)
        for e in self.entries:
            counts[e.res_type] += 1
        return dict(counts)

    def get_resource_type_sizes(self) -> dict[int, int]:
        """Sum file sizes per resource type."""
        sizes = defaultdict(int)
        for e in self.entries:
            sizes[e.res_type] += e.file_size
        return dict(sizes)

    @classmethod
    def type_name(cls, type_id: int) -> str:
        return cls.TYPE_NAMES.get(type_id, f"0x{type_id:08X}")

    @staticmethod
    def quick_validate(filepath) -> tuple[bool, str]:
        """Fast check: is this a valid DBPF file? Returns (valid, reason)."""
        try:
            with open(filepath, "rb") as f:
                header = f.read(12)
            if len(header) < 12:
                return False, "File too small"
            if header[0:4] != b"DBPF":
                return False, f"Bad magic: {header[0:4]!r}"
            major = struct.unpack_from("<I", header, 4)[0]
            if major != 2:
                return False, f"Unsupported version {major}"
            return True, "OK"
        except Exception as e:
            return False, str(e)


# ═══════════════════════════════════════════════════════════════════════════════
# KNOWN BROKEN MOD DATABASE
# ═══════════════════════════════════════════════════════════════════════════════

KNOWN_BROKEN_MODS = [
    # ── Jan 2026 patch (1.121) breakages ──
    {"mod_name": "Alchemist Career", "creator": "BosseladyTV", "issue": "Broken by 1.121 UI changes", "fix": "Await update", "risk": "High", "symptom": "Career panel crash"},
    {"mod_name": "Circus Career", "creator": "BosseladyTV", "issue": "Broken by 1.121 UI changes", "fix": "Await update", "risk": "High", "symptom": "Career panel crash"},
    {"mod_name": "Full-Time Careers Bundle", "creator": "BosseladyTV", "issue": "Broken by 1.121 UI changes", "fix": "Await update", "risk": "High", "symptom": "Career load fail"},
    {"mod_name": "Nobility Career", "creator": "BosseladyTV", "issue": "Broken by 1.121 UI changes", "fix": "Await update", "risk": "High", "symptom": "Career panel crash"},
    {"mod_name": "Paranormal Career", "creator": "BosseladyTV", "issue": "Broken by 1.121 UI changes", "fix": "Await update", "risk": "High", "symptom": "Career load fail"},
    {"mod_name": "Sports Career (TS1/TS3)", "creator": "BosseladyTV", "issue": "Broken by 1.121 UI changes", "fix": "Await update", "risk": "High", "symptom": "Career panel crash"},
    {"mod_name": "Traveling Salesman Career", "creator": "BosseladyTV", "issue": "Broken by 1.121 UI changes", "fix": "Await update", "risk": "High", "symptom": "Career load fail"},
    {"mod_name": "Set Filters on Paintings by Reference", "creator": "Scumbumbo/LittleMsSam", "issue": "Broken script injection", "fix": "Remove or update via LittleMsSam", "risk": "Medium", "symptom": "Painting interaction fail"},
    {"mod_name": "Color Sliders", "creator": "thepancake1/MizoreYukii", "issue": "UI exception on lot save", "fix": "Remove", "risk": "High", "symptom": "CAS/Library crash"},
    {"mod_name": "Appliance Insurance", "creator": "NickyClem", "issue": "Obsolete (replaced by SimSure)", "fix": "Remove", "risk": "Medium", "symptom": "Insurance overlap"},
    {"mod_name": "Plumbing Insurance", "creator": "NickyClem", "issue": "Obsolete (replaced by SimSure)", "fix": "Remove", "risk": "Medium", "symptom": "Insurance overlap"},
    {"mod_name": "Gadget Cover", "creator": "NickyClem", "issue": "Obsolete (replaced by SimSure)", "fix": "Remove", "risk": "Medium", "symptom": "Insurance overlap"},
    {"mod_name": "DSA_BugFix_FightAvailability", "creator": "tyjokr", "issue": "Obsolete after social autonomy fixes", "fix": "Remove", "risk": "Low", "symptom": "Fight autonomy glitch"},
    # ── Ongoing / legacy broken ──
    {"mod_name": "TherapyCenter", "creator": "Unknown", "issue": "HasTunableReference import error post-1.120", "fix": "Remove or update", "risk": "High", "symptom": "Game load crash"},
    {"mod_name": "Knitting Unlocks Bug Fix", "creator": "KaylaDot", "issue": "Obsolete after 1.120 mini-patch", "fix": "Remove", "risk": "High", "symptom": "Crafting crash"},
    {"mod_name": "Bouncer Fix", "creator": "Kuttoe", "issue": "Club crashes post-1.121 UI update", "fix": "Remove (obsolete)", "risk": "High", "symptom": "Venue load fail"},
    {"mod_name": "UI Cheats Extension", "creator": "weerbesu", "issue": "Gallery UI errors if outdated", "fix": "Update to latest", "risk": "Medium", "symptom": "Gallery/UI error"},
    {"mod_name": "More CAS Columns", "creator": "weerbesu", "issue": "CAS layout issues possible", "fix": "Update to latest", "risk": "Low", "symptom": "CAS layout bug"},
    {"mod_name": "Simulation Lag Fix", "creator": "Various", "issue": "Obsolete in 1.120+", "fix": "Remove", "risk": "Medium", "symptom": "Simulation conflict"},
    {"mod_name": "Better BuildBuy", "creator": "TwistedMexi", "issue": "Requires update for 1.120+", "fix": "Update", "risk": "Low", "symptom": "Build mode crash"},
    {"mod_name": "TOOL Mod", "creator": "TwistedMexi", "issue": "Requires update for 1.120+", "fix": "Update", "risk": "Low", "symptom": "Object placement error"},
]

# Filename fragments → known broken mod (for fuzzy matching)
# Key: lowercase substring that must appear in filename (no spaces/underscores)
# Value: index into KNOWN_BROKEN_MODS
_BROKEN_FILENAME_MAP = {}
for _i, _mod in enumerate(KNOWN_BROKEN_MODS):
    _key = _mod["mod_name"].lower().replace(" ", "").replace("-", "").replace("_", "")
    if len(_key) >= 8:  # Only match on names long enough to be unique
        _BROKEN_FILENAME_MAP[_key] = _i


# ═══════════════════════════════════════════════════════════════════════════════
# SCRIPT MOD VERSION REGISTRY
# ═══════════════════════════════════════════════════════════════════════════════
# Maps known script mods to version detection patterns and safe version info.
#
# filename_match: list of lowercase substrings — if ANY match the filename,
#                 this entry is considered a candidate.
# version_patterns: list of regex patterns to extract version from tuning XML,
#                   script source, or filenames. First match wins.
# min_safe_version: tuple of ints — minimum version considered safe for 1.121+
# check_scripts: if True, also look inside .ts4script zips for version strings
# notes: human-readable context

KNOWN_SCRIPT_MODS = [
    {
        "name": "MC Command Center",
        "creator": "Deaderpool",
        "filename_match": ["mc_cmd_center", "mccc", "mc_command"],
        "version_patterns": [
            r"mc_cmd_center[_\-]?v?(\d{4})[._](\d+)[._](\d+)",
            r"MCCC[_ ]?v?(\d{4})[._](\d+)[._](\d+)",
            r"McCmdCenter.*?(\d{4})[_.](\d+)[_.](\d+)",
            r"AllModules[_]?(\d{4})[_.](\d+)[_.](\d+)",
            r"version.*?(\d{4})[._](\d+)[._](\d+)",
        ],
        "min_safe_version": (2025, 7, 0),
        "check_scripts": True,
        "notes": "Keep updated — folder name may not reflect actual version",
    },
    {
        "name": "UI Cheats Extension",
        "creator": "weerbesu",
        "filename_match": ["ui_cheats_extension", "uicheats"],
        "version_patterns": [
            r"ui_cheats_extension[_\-]?v?(\d+)[._](\d+)(?:[._](\d+))?",
            r"version.*?(\d+)[._](\d+)[._](\d+)",
            r"UICheat.*?(\d+)[._](\d+)(?:[._](\d+))?",
        ],
        "min_safe_version": (1, 43, 0),
        "check_scripts": True,
        "notes": "Gallery UI errors on pre-1.43",
    },
    {
        "name": "Better BuildBuy",
        "creator": "TwistedMexi",
        "filename_match": ["betterbuildbuy", "tmex-betterbuildbuy"],
        "version_patterns": [
            r"betterbuildbuy[_\-]?v?(\d+)[._](\d+)(?:[._](\d+))?",
            r"tmex.*?build.*?(\d+)[._](\d+)(?:[._](\d+))?",
            r"version.*?(\d+)[._](\d+)(?:[._](\d+))?",
        ],
        "min_safe_version": (3, 7, 0),
        "check_scripts": True,
        "notes": "Build mode crash if pre-3.7",
    },
    {
        "name": "TOOL Mod",
        "creator": "TwistedMexi",
        "filename_match": ["tmex-tool", "tool_mod"],
        "version_patterns": [
            r"tool[_\-]?v?(\d+)[._](\d+)(?:[._](\d+))?",
            r"version.*?(\d+)[._](\d+)(?:[._](\d+))?",
        ],
        "min_safe_version": (8, 2, 0),
        "check_scripts": True,
        "notes": "Object placement errors if outdated",
    },
    {
        "name": "WickedWhims",
        "creator": "TurboDriver",
        "filename_match": ["turbodriver_wickedwhims"],
        "exclude_substrings": [
            "animation", "lupobianco", "beebavel", "exotica", "noir",
            "language", "vulgar", "startingpack", "perfectlyround",
            "smoothvagina", "werewolves", "futanari", "penis",
        ],
        "version_patterns": [
            r"wickedwhims[_\-]?v?(\d+)(?:[._](\d+))?(?:[._](\d+))?",
            r"version.*?(\d+)(?:[._](\d+))?(?:[._](\d+))?",
        ],
        "min_safe_version": (187, 0, 0),
        "check_scripts": True,
        "notes": "Major mod — keep updated",
    },
    {
        "name": "More CAS Columns",
        "creator": "weerbesu",
        "filename_match": ["more_cas_columns", "morecascolumns"],
        "version_patterns": [
            r"more_cas_columns.*?(\d+)[._]?(\d+)?(?:[._](\d+))?",
            r"version.*?(\d+)[._](\d+)(?:[._](\d+))?",
        ],
        "min_safe_version": (2, 5, 0),
        "check_scripts": True,
        "notes": "CAS layout issues if outdated",
    },
    {
        "name": "Basemental Drugs",
        "creator": "Basemental",
        "filename_match": ["basementaldrugs", "basemental_drugs"],
        "version_patterns": [
            r"basemental.*?drugs.*?v?(\d+)[._](\d+)(?:[._](\d+))?",
            r"version.*?(\d+)[._](\d+)(?:[._](\d+))?",
        ],
        "min_safe_version": (8, 0, 0),
        "check_scripts": True,
        "notes": "Script mod — keep updated with game patches",
    },
    {
        "name": "Lot51 Core Library",
        "creator": "Lot51",
        "filename_match": ["lot51_core", "lot51core"],
        "version_patterns": [
            r"lot51.*?core.*?v?(\d+)[._](\d+)(?:[._](\d+))?",
            r"version.*?(\d+)[._](\d+)(?:[._](\d+))?",
        ],
        "min_safe_version": (1, 38, 0),
        "check_scripts": True,
        "notes": "Core library used by multiple Lot51 mods",
    },
    {
        "name": "Nisa's Wicked Perversions",
        "creator": "NisaK",
        "filename_match": ["nisawickedpervs", "nisa_wicked", "wickedperversions"],
        "version_patterns": [
            r"nisa.*?v?(\d+)[._](\d+)(?:[._](\d+))?",
            r"version.*?(\d+)[._](\d+)(?:[._](\d+))?",
        ],
        "min_safe_version": (2, 7, 0),
        "check_scripts": True,
        "notes": "Script mod — keep updated",
    },
    {
        "name": "Sacrificial Mods",
        "creator": "Sacrificial",
        "filename_match": ["sacrificial_", "life_tragedies", "extreme_violence"],
        "version_patterns": [
            r"sacrificial.*?v?(\d+)[._](\d+)(?:[._](\d+))?",
            r"version.*?(\d+)[._](\d+)(?:[._](\d+))?",
        ],
        "min_safe_version": (1, 0, 0),
        "check_scripts": True,
        "notes": "Script mods — check for latest",
    },
]


# ═══════════════════════════════════════════════════════════════════════════════
# CC METADATA IDENTIFICATION
# ═══════════════════════════════════════════════════════════════════════════════

# Known creator identifiers (lowercase prefix/substring → display name)
KNOWN_CREATORS = {
    "turbodriver": "TurboDriver",
    "basemental": "Basemental",
    "nisak": "NisaK",
    "nisa_": "NisaK",
    "pandasama": "PandaSama",
    "mccc": "Deaderpool",
    "mc_cmd": "Deaderpool",
    "deaderpool": "Deaderpool",
    "ww_": "WickedWhims Community",
    "wickedwhims": "TurboDriver",
    "aggressivekitty": "Aggressive Kitty",
    "felixandre": "Felixandre",
    "harrie": "Harrie",
    "pierisim": "Pierisim",
    "slephora": "Slephora",
    "obscurus": "Obscurus Sims",
    "doux": "Doux",
    "onyxsims": "OnyxSims",
    "babybeesims": "BabyBeeSims",
    "rubyred": "RubyRed",
    "littlemssam": "LittleMsSam",
    "kuttoe": "Kuttoe",
    "lot51": "Lot51",
    "ravasheen": "Ravasheen",
    "ilkavelle": "IlkaVelle",
    "simrealist": "SimRealist",
    "adeepindigo": "ADeepIndigo",
    "zerbu": "Zerbu",
    "kawaiistacie": "KawaiiStacie",
    "sacrificial": "Sacrificial",
    "lumpinou": "Lumpinou",
    "tmex": "TwistedMexi",
    "bienchen": "Bienchen",
    "roburky": "Roburky",
    "camuflaje": "Camuflaje",
    "greenllamas": "GreenLlamas",
    "arethabee": "ArethaBee",
    "coloresurbanos": "ColoresUrbanos",
    "simcelebrity": "SimCelebrity",
    "snootysims": "SnootySims",
    "teknikah": "Teknikah",
    "weerbesu": "weerbesu",
}


def identify_cc_type(filename_lower: str, suffix: str, type_counts: dict | None = None) -> str:
    """Determine CC type from filename and optionally from parsed resource types."""
    # If we have DBPF resource data, use it for better classification
    if type_counts:
        has_cas = type_counts.get(DBPFParser.CAS_PART, 0) > 0
        has_geom = type_counts.get(DBPFParser.GEOM, 0) > 0
        has_tuning = type_counts.get(DBPFParser.TUNING_XML, 0) > 0
        has_clip = type_counts.get(DBPFParser.CLIP, 0) > 0
        has_objdef = type_counts.get(DBPFParser.OBJDEF, 0) > 0

        if has_cas and has_geom:
            # CAS CC — try to narrow down from filename
            for hint, label in [
                ("hair", "Hair CC"), ("hairstyle", "Hair CC"),
                ("skin", "Skin/Overlay CC"), ("overlay", "Skin/Overlay CC"),
                ("makeup", "Makeup CC"), ("lipstick", "Makeup CC"),
                ("eyeshadow", "Makeup CC"), ("blush", "Makeup CC"),
                ("clothing", "Clothing CC"), ("outfit", "Clothing CC"),
                ("top_", "Clothing CC"), ("bottom_", "Clothing CC"),
                ("dress", "Clothing CC"), ("shoes", "Clothing CC"),
                ("accessory", "Accessory CC"), ("earring", "Accessory CC"),
                ("necklace", "Accessory CC"), ("glasses", "Accessory CC"),
                ("hat", "Accessory CC"), ("ring", "Accessory CC"),
            ]:
                if hint in filename_lower:
                    return label
            return "CAS CC"
        elif has_cas and not has_geom:
            return "CAS Override"
        elif has_objdef or (has_geom and not has_cas):
            for hint, label in [
                ("furniture", "Furniture/Build CC"), ("counter", "Furniture/Build CC"),
                ("chair", "Furniture/Build CC"), ("table", "Furniture/Build CC"),
                ("sofa", "Furniture/Build CC"), ("bed", "Furniture/Build CC"),
                ("kitchen", "Furniture/Build CC"), ("wall", "Build CC"),
                ("floor", "Build CC"), ("roof", "Build CC"), ("window", "Build CC"),
                ("door", "Build CC"), ("fence", "Build CC"),
            ]:
                if hint in filename_lower:
                    return label
            return "Object CC"
        elif has_clip:
            return "Animation/Pose"
        elif has_tuning and not has_geom and not has_cas:
            return "Tuning Mod"

    # Filename-based fallback
    if suffix == ".ts4script":
        return "Script Mod"
    for hint, label in [
        ("hair", "Hair CC"), ("skin", "Skin CC"), ("overlay", "Skin CC"),
        ("makeup", "Makeup CC"), ("clothing", "Clothing CC"),
        ("furniture", "Furniture CC"), ("trait", "Trait/Buff Mod"),
        ("buff", "Trait/Buff Mod"), ("animation", "Animation"),
        ("pose", "Pose Pack"),
    ]:
        if hint in filename_lower:
            return label
    return "Package"


def identify_creator(filename: str) -> str:
    """Try to identify the CC creator from the filename."""
    fn_lower = filename.lower()
    # Check known creators
    for key, name in KNOWN_CREATORS.items():
        if key in fn_lower:
            return name
    # Heuristic: many CC files use "Creator_Description" format
    # Try first segment before underscore if it's a plausible name
    if "_" in filename:
        first = filename.split("_")[0].strip()
        # Filter out things that are clearly not creator names
        if (len(first) >= 3 and not first[0].isdigit()
                and first.lower() not in ("the", "new", "old", "fix", "mod", "sim", "cas", "my")):
            return first
    # Check for [Creator] bracket format
    bracket = re.match(r"\[([^\]]+)\]", filename)
    if bracket:
        return bracket.group(1).strip()
    return "Unknown"


# ═══════════════════════════════════════════════════════════════════════════════
# MOD ANALYZER — Core analysis engine
# ═══════════════════════════════════════════════════════════════════════════════

def _hash_file(filepath) -> tuple[str | None, str, Path]:
    """Worker function for parallel hashing. Returns (hash, clean_name, path)."""
    try:
        md5 = hashlib.md5()
        with open(filepath, "rb") as f:
            while chunk := f.read(65536):
                md5.update(chunk)
        base = filepath.stem.lower()
        clean = re.sub(r"[-_]?v?\d+[\d.]*$", "", base)
        clean = re.sub(r"[-_](updated?|new|fixed?|patch|old|backup)$", "", clean)
        return md5.hexdigest(), clean, filepath
    except Exception:
        return None, "", filepath


class ModAnalyzer:
    """Main analysis engine for Sims 4 mod folders."""

    def __init__(self, mods_path: str, log_path: str | None = None,
                 modlist_path: str | None = None):
        self.mods_path = Path(mods_path)
        self.log_path = Path(log_path) if log_path else None
        self.modlist_path = Path(modlist_path) if modlist_path else None
        self.packages: list[Path] = []
        self.scripts: list[Path] = []
        self.total_size_bytes = 0
        self.issues: list[dict] = []   # All detected issues (for CSV + report)
        self.dbpf_cache: dict[Path, DBPFParser] = {}  # Parsed package cache
        self.parse_failures: list[tuple[Path, str]] = []
        self.modlist_db: list[dict] = []       # Parsed Scarlet mod checker entries
        self.modlist_matches: list[dict] = []  # Cross-reference results

    # ── Helpers ──

    def _add_issue(self, issue_type: str, filepath: Path | str, risk: str,
                   symptom: str, notes: str, size: int = 0, date: str = ""):
        self.issues.append({
            "Type": issue_type,
            "File": str(filepath),
            "Size": size,
            "Date": date,
            "Notes": notes,
            "Risk": risk,
            "Symptom": symptom,
        })

    def _file_date(self, path: Path) -> str:
        try:
            return datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d")
        except Exception:
            return ""

    def _file_size(self, path: Path) -> int:
        try:
            return path.stat().st_size
        except Exception:
            return 0

    def _parse_package(self, filepath: Path) -> DBPFParser | None:
        """Parse a package, using cache. Returns None on failure."""
        if filepath in self.dbpf_cache:
            return self.dbpf_cache[filepath]
        parser = DBPFParser(filepath)
        try:
            parser.parse()
            self.dbpf_cache[filepath] = parser
            return parser
        except DBPFError as e:
            self.parse_failures.append((filepath, str(e)))
            return None

    @property
    def total_size_mb(self) -> float:
        return self.total_size_bytes / (1024 * 1024)

    # ── 1. Scan Mods Folder ──

    def scan_mods(self) -> bool:
        """Walk the mods directory and collect .package and .ts4script files."""
        print(f"Scanning: {self.mods_path}")
        if not self.mods_path.exists():
            print(f"  ERROR: Folder not found")
            return False

        for root, dirs, files in os.walk(self.mods_path):
            for fname in files:
                ext = fname.lower().rsplit(".", 1)[-1] if "." in fname else ""
                if ext == "package":
                    fp = Path(root) / fname
                    self.packages.append(fp)
                    self.total_size_bytes += self._file_size(fp)
                elif ext == "ts4script":
                    fp = Path(root) / fname
                    self.scripts.append(fp)
                    self.total_size_bytes += self._file_size(fp)

        print(f"  Found {len(self.packages):,} packages, {len(self.scripts):,} scripts")
        print(f"  Total size: {self.total_size_mb:.1f} MB ({self.total_size_mb/1024:.2f} GB)")
        return True

    # ── 2. Validate DBPF Integrity ──

    def check_integrity(self) -> int:
        """Validate every .package has a proper DBPF header and parseable index."""
        print("\n" + "=" * 70)
        print("PACKAGE INTEGRITY CHECK")
        print("=" * 70)

        corrupt = []
        empty = []
        tiny = []

        for fp in tqdm(self.packages, desc="Validating headers"):
            size = self._file_size(fp)
            if size == 0:
                empty.append(fp)
                continue
            if size < 96:
                tiny.append((fp, size))
                continue

            valid, reason = DBPFParser.quick_validate(fp)
            if not valid:
                corrupt.append((fp, size, reason))

        if empty:
            print(f"\n  [!] EMPTY FILES ({len(empty)}):")
            for fp in empty:
                rel = fp.relative_to(self.mods_path)
                print(f"      {rel}")
                self._add_issue("integrity", fp, "High", "Empty/corrupt",
                                "0 bytes — placeholder or failed download",
                                0, self._file_date(fp))

        if tiny:
            print(f"\n  [!] SUSPICIOUSLY SMALL ({len(tiny)}):")
            for fp, sz in tiny:
                rel = fp.relative_to(self.mods_path)
                print(f"      {rel} — {sz} bytes")
                self._add_issue("integrity", fp, "Medium", "Possible corruption",
                                f"Only {sz} bytes — too small for valid package",
                                sz, self._file_date(fp))

        if corrupt:
            print(f"\n  [!] CORRUPT/INVALID PACKAGES ({len(corrupt)}):")
            for fp, sz, reason in corrupt:
                rel = fp.relative_to(self.mods_path)
                print(f"      {rel} — {reason}")
                self._add_issue("integrity", fp, "High", "Invalid package",
                                f"DBPF parse failed: {reason}",
                                sz, self._file_date(fp))

        total = len(empty) + len(tiny) + len(corrupt)
        if total == 0:
            print("\n  All packages passed integrity check.")
        else:
            print(f"\n  {total} integrity issues found.")
        return total

    # ── 3. Detect Duplicates ──

    def detect_duplicates(self, parallel: bool = True) -> int:
        """Find exact duplicate files (by hash) and similar filenames."""
        print("\n" + "=" * 70)
        print("DUPLICATE DETECTION")
        print("=" * 70)

        all_files = self.packages + self.scripts
        hash_map = defaultdict(list)
        name_map = defaultdict(list)

        print(f"  Hashing {len(all_files):,} files...")
        if parallel and len(all_files) > 100:
            with Pool(min(cpu_count(), 8)) as pool:
                results = list(tqdm(
                    pool.imap(_hash_file, all_files, chunksize=200),
                    total=len(all_files), desc="  Hashing (parallel)"))
        else:
            results = [_hash_file(f) for f in tqdm(all_files, desc="  Hashing")]

        for file_hash, clean_name, filepath in results:
            if file_hash:
                hash_map[file_hash].append(filepath)
            if clean_name:
                name_map[clean_name].append(filepath)

        exact = {h: paths for h, paths in hash_map.items() if len(paths) > 1}
        similar = {n: paths for n, paths in name_map.items()
                   if len(paths) > 1 and n.strip() and len(n) >= 4}

        if exact:
            print(f"\n  EXACT DUPLICATES — {len(exact)} sets:")
            for i, (h, paths) in enumerate(exact.items(), 1):
                sz = self._file_size(paths[0])
                print(f"\n    Set #{i} — {len(paths)} identical files ({sz:,} bytes):")
                for p in sorted(paths):
                    rel = p.relative_to(self.mods_path)
                    print(f"      [X] {rel}")
                    self._add_issue("duplicate_exact", p, "Medium",
                                    "Wasted space / potential conflict",
                                    f"Hash: {h[:12]}... ({len(paths)} copies)",
                                    sz, self._file_date(p))
                print(f"    → Keep only ONE, delete {len(paths)-1}")
        else:
            print("\n  No exact duplicates found.")

        # Filter out sets already covered by exact dupes
        exact_sets = {frozenset(str(p) for p in paths) for paths in exact.values()}
        filtered_similar = {}
        for name, paths in similar.items():
            path_strs = frozenset(str(p) for p in paths)
            if not any(path_strs <= es for es in exact_sets):
                filtered_similar[name] = paths

        if filtered_similar:
            print(f"\n  SIMILAR FILENAMES — {len(filtered_similar)} groups:")
            for i, (name, paths) in enumerate(sorted(filtered_similar.items())[:30], 1):
                print(f"\n    Group #{i}: '{name}' — {len(paths)} files:")
                for p in sorted(paths):
                    rel = p.relative_to(self.mods_path)
                    sz = self._file_size(p)
                    print(f"      [!] {rel} ({sz:,} bytes, {self._file_date(p)})")
                print(f"    → Keep the newest/largest version")

        return len(exact) + len(filtered_similar)

    # ── 4. Detect Outdated Mods ──

    def detect_outdated(self) -> int:
        """Flag files by modification date. Note: mtime is unreliable for CC
        downloaded from file-sharing sites (which often reset timestamps), but
        it's the best heuristic available without parsing internal version data."""
        print("\n" + "=" * 70)
        print("OUTDATED MOD DETECTION (by file modification date)")
        print("=" * 70)
        print("  Note: File dates can be inaccurate if mods were redownloaded")
        print("  or moved. Treat as a rough guide, not definitive.")

        now = datetime.now().timestamp()
        one_year = now - (365 * 86400)
        two_years = now - (730 * 86400)

        old_2y = []
        old_1y = []

        for fp in self.packages + self.scripts:
            try:
                mtime = fp.stat().st_mtime
                if mtime < two_years:
                    old_2y.append((fp, datetime.fromtimestamp(mtime)))
                elif mtime < one_year:
                    old_1y.append((fp, datetime.fromtimestamp(mtime)))
            except Exception:
                pass

        if old_2y:
            old_2y.sort(key=lambda x: x[1])
            print(f"\n  CRITICALLY OUTDATED (2+ years) — {len(old_2y)} files:")
            for fp, dt in old_2y[:50]:
                rel = fp.relative_to(self.mods_path)
                age = (now - dt.timestamp()) / (365 * 86400)
                print(f"    [X] {rel}  ({dt.strftime('%Y-%m-%d')}, {age:.1f}y old)")
                self._add_issue("outdated_critical", fp, "High",
                                "Likely incompatible",
                                f"Last modified {dt.strftime('%Y-%m-%d')} ({age:.1f} years)",
                                self._file_size(fp), dt.strftime("%Y-%m-%d"))
            if len(old_2y) > 50:
                print(f"    ... and {len(old_2y)-50} more")

        if old_1y:
            old_1y.sort(key=lambda x: x[1])
            print(f"\n  OUTDATED (1-2 years) — {len(old_1y)} files:")
            for fp, dt in old_1y[:30]:
                rel = fp.relative_to(self.mods_path)
                print(f"    [!] {rel}  ({dt.strftime('%Y-%m-%d')})")
                self._add_issue("outdated", fp, "Medium", "May need update",
                                f"Last modified {dt.strftime('%Y-%m-%d')}",
                                self._file_size(fp), dt.strftime("%Y-%m-%d"))
            if len(old_1y) > 30:
                print(f"    ... and {len(old_1y)-30} more")

        if not old_2y and not old_1y:
            print("\n  No outdated files detected.")
        return len(old_2y) + len(old_1y)

    # ── 5. Detect Known Broken Mods ──

    def detect_broken_mods(self) -> int:
        """Match installed mods against the known broken database."""
        print("\n" + "=" * 70)
        print("KNOWN BROKEN MOD DETECTION")
        print("=" * 70)

        # Also load local JSON database if it exists
        local_db = self._load_local_db()
        all_known = KNOWN_BROKEN_MODS + local_db

        matches = []
        all_files = self.packages + self.scripts

        for fp in all_files:
            fn_clean = (fp.stem.lower()
                        .replace(" ", "").replace("-", "").replace("_", ""))

            for mod_info in all_known:
                mod_key = (mod_info["mod_name"].lower()
                           .replace(" ", "").replace("-", "").replace("_", ""))
                # Require minimum match length to avoid false positives
                if len(mod_key) >= 8 and mod_key in fn_clean:
                    matches.append((fp, mod_info))
                    break
                # Also check with word boundaries for shorter names
                elif len(mod_key) >= 4:
                    # Check if the filename essentially IS the mod name
                    if fn_clean == mod_key or fn_clean.startswith(mod_key):
                        matches.append((fp, mod_info))
                        break

        if matches:
            print(f"\n  MATCHED {len(matches)} potentially broken mods:")
            for fp, info in matches:
                rel = fp.relative_to(self.mods_path)
                print(f"\n    [{info['risk'].upper()}] {rel}")
                print(f"      Mod: {info['mod_name']} by {info['creator']}")
                print(f"      Issue: {info['issue']}")
                print(f"      Fix: {info['fix']}")
                self._add_issue("known_broken", fp, info["risk"],
                                info["symptom"],
                                f"{info['mod_name']} — {info['issue']} | Fix: {info['fix']}",
                                self._file_size(fp), self._file_date(fp))
        else:
            print("\n  No known broken mods detected.")
        return len(matches)

    def _load_local_db(self) -> list[dict]:
        """Load additional broken mod entries from local JSON if present."""
        db_path = self.mods_path / "broken_cc_hashes.json"
        if not db_path.exists():
            return []
        try:
            with open(db_path, "r") as f:
                data = json.load(f)
            entries = data.get("known_broken", [])
            # Validate structure
            valid = []
            for e in entries:
                if isinstance(e, dict) and "mod_name" in e:
                    valid.append({
                        "mod_name": e.get("mod_name", ""),
                        "creator": e.get("creator", "Unknown"),
                        "issue": e.get("issue", "Unknown issue"),
                        "fix": e.get("fix", "Check for update"),
                        "risk": e.get("risk", "Medium"),
                        "symptom": e.get("symptom", "Unknown"),
                    })
            return valid
        except Exception:
            return []

    # ── 5b. Load Scarlet Mod Checker List ──

    def load_modlist(self) -> int:
        """Load the Scarlet Realm / Mod Checker CSV into a lookup database."""
        if not self.modlist_path or not self.modlist_path.exists():
            return 0

        print("\n" + "=" * 70)
        print("LOADING MOD CHECKER LIST")
        print("=" * 70)
        print(f"  File: {self.modlist_path.name}")

        try:
            with open(self.modlist_path, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    mod_name = row.get("Mod Name", "").strip()
                    if not mod_name:
                        continue
                    self.modlist_db.append({
                        "mod_name": mod_name,
                        "creator": row.get("Creator", "").strip(),
                        "link": row.get("Link", "").strip(),
                        "patch_status": row.get("Patch Status", "").strip(),
                        "last_update": row.get("Last Known Update", "").strip(),
                        "status_date": row.get("Last Status Change (YYYY-MM-DD)", "").strip(),
                        "notes": row.get("Notes", "").strip(),
                        "info_link": row.get("Additional Info Link", "").strip(),
                    })
        except Exception as e:
            print(f"  [Error] Could not parse mod checker CSV: {e}")
            return 0

        # Build lookup keys: normalized mod names for fuzzy matching
        # We store a clean version (lowercase, no spaces/punctuation) for matching
        for entry in self.modlist_db:
            raw = entry["mod_name"]
            clean = re.sub(r"[^a-z0-9]", "", raw.lower())
            entry["_match_key"] = clean
            # Also store individual significant words (3+ chars) for partial matching
            words = [w.lower() for w in re.findall(r"[A-Za-z0-9]{3,}", raw)]
            entry["_match_words"] = words

        statuses = {}
        for e in self.modlist_db:
            s = e["patch_status"]
            if s:
                statuses[s] = statuses.get(s, 0) + 1

        print(f"  Loaded {len(self.modlist_db)} mods from checker list")
        print(f"  Status breakdown:")
        for status, count in sorted(statuses.items(), key=lambda x: -x[1]):
            print(f"    {status:<20} {count:>5}")

        return len(self.modlist_db)

    # ── 5c. Cross-Reference Installed Mods Against Mod Checker List ──

    def cross_reference_modlist(self) -> int:
        """Match installed mods against the Scarlet Mod Checker list.
        Uses strict matching to avoid false positives — only matches when
        we're confident the mod in your folder IS the mod on the list."""

        if not self.modlist_db:
            return 0

        print("\n" + "=" * 70)
        print("CROSS-REFERENCING WITH MOD CHECKER LIST")
        print("=" * 70)

        all_files = self.packages + self.scripts
        matched = 0
        broken_count = 0
        updated_count = 0
        issue_count = 0

        # Pre-build a fast lookup: clean key → list of modlist entries
        key_lookup = defaultdict(list)
        for entry in self.modlist_db:
            key_lookup[entry["_match_key"]].append(entry)

        # Names too generic to match by substring — would cause false positives
        GENERIC_NAMES = {
            "smallmods", "career", "trait", "traits", "mod", "mods",
            "overrides", "override", "fix", "patch", "update", "addon",
            "activities", "custom", "more", "better", "new", "extra",
        }

        # Also build creator-based lookup for tighter matching
        creator_lookup = defaultdict(list)
        for entry in self.modlist_db:
            creator_clean = re.sub(r"[^a-z0-9]", "", entry["creator"].lower().split("(")[0].strip())
            if creator_clean:
                creator_lookup[creator_clean].append(entry)

        for fp in tqdm(all_files, desc="  Cross-referencing mods"):
            filename = fp.stem
            fn_clean = re.sub(r"[^a-z0-9]", "", filename.lower())

            # Skip very short filenames (likely just numbers or fragments)
            if len(fn_clean) < 6:
                continue

            best_match = None
            best_score = 0

            # ── Strategy 1: Exact key match (cleaned mod name == cleaned filename)
            if fn_clean in key_lookup:
                best_match = key_lookup[fn_clean][0]
                best_score = 100

            # ── Strategy 2: Creator + mod name in filename
            # If filename contains a known creator name, match against
            # that creator's mods only — much more accurate
            if not best_match:
                for creator_clean, entries in creator_lookup.items():
                    if len(creator_clean) < 4:
                        continue
                    if creator_clean in fn_clean:
                        # Found creator in filename — now check mod names
                        # Strip creator from filename for cleaner comparison
                        fn_no_creator = fn_clean.replace(creator_clean, "", 1)

                        for entry in entries:
                            mk = entry["_match_key"]
                            if len(mk) < 6:
                                continue

                            # Check if mod name key is fully in filename
                            if mk in fn_clean:
                                score = 95
                                if score > best_score:
                                    best_match = entry
                                    best_score = score
                                    break

                            # Creator confirmed: check if filename remainder is
                            # a prefix/substring of the mod key (handles cases like
                            # "LittleMsSam_21DaysSeasons" → "21 Days Seasons instead of 28 Days")
                            if len(fn_no_creator) >= 8:
                                if fn_no_creator in mk or mk.startswith(fn_no_creator):
                                    score = 90
                                    if score > best_score:
                                        best_match = entry
                                        best_score = score

                            # Check significant words from the mod name (5+ char words)
                            sig_words = [w for w in entry["_match_words"] if len(w) >= 5]
                            if len(sig_words) >= 2:
                                hits = sum(1 for w in sig_words if w in fn_clean)
                                # Need majority of significant words to match
                                if hits >= 2 and hits >= len(sig_words) * 0.6:
                                    score = 85
                                    if score > best_score:
                                        best_match = entry
                                        best_score = score

            # ── Strategy 3: Strict substring — modlist key must be a LARGE
            # portion of the filename (not just a tiny fragment)
            if not best_match:
                for entry in self.modlist_db:
                    mk = entry["_match_key"]
                    # Require long match keys to avoid "career" matching everything
                    if len(mk) < 12:
                        continue
                    # Skip generic mod names that would match too broadly
                    if mk in GENERIC_NAMES:
                        continue

                    if mk in fn_clean:
                        # The match key should be at least 60% of the filename
                        overlap = len(mk) / max(len(fn_clean), 1)
                        if overlap >= 0.55:
                            score = overlap * 100
                            if score > best_score:
                                best_match = entry
                                best_score = score

                    elif fn_clean in mk and len(fn_clean) >= 12:
                        overlap = len(fn_clean) / max(len(mk), 1)
                        if overlap >= 0.55:
                            score = overlap * 95
                            if score > best_score:
                                best_match = entry
                                best_score = score

            # Only accept matches we're confident about (score >= 55)
            # CRITICAL: Reject matches where the file clearly belongs to a
            # different creator than the matched mod on the Scarlet list
            if best_match and best_score >= 55:
                reject = False
                matched_creator_clean = re.sub(
                    r"[^a-z0-9]", "",
                    best_match["creator"].lower().split("(")[0].strip())

                # ── Check 1: Known creator in filename vs matched creator
                for creator_clean in creator_lookup:
                    if len(creator_clean) < 4:
                        continue
                    if creator_clean in fn_clean:
                        if (creator_clean not in matched_creator_clean
                                and matched_creator_clean not in creator_clean):
                            reject = True
                        break

                # ── Check 2: Filename prefix abbreviations
                # Many creators use abbreviations: LBB=LittleBowBub, SS=SomikSeverinka, etc.
                if not reject:
                    CREATOR_ABBREVS = {
                        "lbb": "littlebowbub", "lms": "littlemssam",
                        "rvsn": "ravasheen", "tmex": "twistedmexi",
                        "ks": "kawaiistacie", "td": "turbodriver",
                        "nc": "nickyclem", "kl": "kawaiilogan",
                        "zer": "zerbu", "sac": "sacrificial",
                        "bm": "basemental", "rb": "roburky",
                        "pbs": "polarbearsims", "ics": "icemunmun",
                        "snb": "snootysims", "mei": "meiusto",
                    }
                    fn_prefix = fp.stem.split("_")[0].lower() if "_" in fp.stem else ""
                    fn_prefix_clean = re.sub(r"[^a-z0-9]", "", fn_prefix)
                    expanded = CREATOR_ABBREVS.get(fn_prefix_clean, "")
                    if expanded and len(expanded) >= 4:
                        if (expanded not in matched_creator_clean
                                and matched_creator_clean not in expanded):
                            reject = True

                # ── Check 3: Subfolder name as creator verification
                # If the file is in a subfolder named after a creator, and the
                # matched mod is by a DIFFERENT creator, reject it
                if not reject:
                    try:
                        rel = fp.relative_to(self.mods_path)
                        rel_parts = rel.parts
                        if len(rel_parts) >= 2:
                            subfolder = rel_parts[0].lower()
                            sub_clean = re.sub(r"[^a-z0-9]", "", subfolder)
                            if len(sub_clean) >= 5:
                                # Check if subfolder looks like a creator name
                                # (not a generic category like "mods", "cc", "scripts")
                                generic_folders = {
                                    "mods", "scripts", "cc", "package", "packages",
                                    "tuning", "overrides", "merged", "zmerged",
                                    "bodyslider", "body slider", "hair", "clothing",
                                    "furniture", "build", "cas", "debug", "old",
                                    "backup", "temp", "test", "new", "updated",
                                }
                                if sub_clean not in generic_folders:
                                    if (sub_clean not in matched_creator_clean
                                            and matched_creator_clean not in sub_clean):
                                        # Subfolder is a creator name that doesn't
                                        # match — but only reject if subfolder
                                        # also doesn't match the mod name itself
                                        mod_key = best_match["_match_key"]
                                        if sub_clean not in mod_key:
                                            reject = True
                    except (ValueError, IndexError):
                        pass

                if reject:
                    continue  # Skip this false match
            if best_match and best_score >= 55:
                matched += 1
                status = best_match["patch_status"]
                match_record = {
                    "file": fp,
                    "filename": filename,
                    "matched_mod": best_match["mod_name"],
                    "creator": best_match["creator"],
                    "patch_status": status,
                    "last_update": best_match["last_update"],
                    "status_date": best_match["status_date"],
                    "notes": best_match["notes"],
                    "info_link": best_match["info_link"],
                    "match_score": best_score,
                }
                self.modlist_matches.append(match_record)

                # Add issues for problematic statuses
                if status == "Broken":
                    broken_count += 1
                    self._add_issue("modlist_broken", fp, "High",
                                    f"Broken per Mod Checker ({best_match['status_date']})",
                                    f"{filename} → {best_match['mod_name']} by {best_match['creator']} — "
                                    f"BROKEN | {best_match['notes'][:100]}",
                                    self._file_size(fp), self._file_date(fp))
                elif status == "Obsolete":
                    issue_count += 1
                    self._add_issue("modlist_obsolete", fp, "High",
                                    "Obsolete per Mod Checker",
                                    f"{filename} → {best_match['mod_name']} by {best_match['creator']} — "
                                    f"OBSOLETE | {best_match['notes'][:100]}",
                                    self._file_size(fp), self._file_date(fp))
                elif status == "Minor Issue(s)":
                    issue_count += 1
                    self._add_issue("modlist_minor_issue", fp, "Medium",
                                    "Minor issues per Mod Checker",
                                    f"{filename} → {best_match['mod_name']} by {best_match['creator']} — "
                                    f"{best_match['notes'][:120]}",
                                    self._file_size(fp), self._file_date(fp))
                elif status == "Updated":
                    updated_count += 1
                elif status == "Unknown":
                    self._add_issue("modlist_unknown", fp, "Low",
                                    "Status unknown on Mod Checker",
                                    f"{filename} → {best_match['mod_name']} by {best_match['creator']} — "
                                    f"Status unknown, verify manually",
                                    self._file_size(fp), self._file_date(fp))

        # Print summary
        print(f"\n  CROSS-REFERENCE RESULTS:")
        print(f"  {'─'*50}")
        print(f"  Installed mods matched:  {matched} / {len(all_files)}")
        print(f"  Unmatched (CC/custom):   {len(all_files) - matched}")
        print()

        # Group by status
        status_groups = defaultdict(list)
        for m in self.modlist_matches:
            status_groups[m["patch_status"]].append(m)

        status_colors = {
            "Broken": "[!!!]", "Obsolete": "[!!!]", "Minor Issue(s)": "[!!]",
            "Updated": "[OK]", "No Known Issues": "[OK]", "N/A": "[--]",
            "Unknown": "[??]", "New": "[OK]", "": "[??]",
        }

        for status in ["Broken", "Obsolete", "Minor Issue(s)", "Unknown",
                        "Updated", "No Known Issues", "N/A", "New", ""]:
            mods = status_groups.get(status, [])
            if not mods:
                continue
            icon = status_colors.get(status, "[??]")
            label = status if status else "No Status"
            print(f"  {icon} {label}: {len(mods)}")

            # Show details for problematic ones
            if status in ("Broken", "Obsolete", "Minor Issue(s)"):
                for m in mods[:15]:
                    print(f"        {m['filename']}")
                    print(f"          → {m['matched_mod']} by {m['creator']}")
                    if m["notes"]:
                        print(f"          Note: {m['notes'][:80]}")
                if len(mods) > 15:
                    print(f"        ... and {len(mods)-15} more")

        return matched

    # ── 6. TGI Conflict Detection (Real DBPF Parsing) ──

    def detect_tuning_conflicts(self) -> int:
        """Find packages that override the same tuning resources by checking
        for duplicate (Type, Instance) pairs across packages.

        This uses the DBPF index — no decompression needed, so it's fast."""
        print("\n" + "=" * 70)
        print("TUNING CONFLICT DETECTION (DBPF Index)")
        print("=" * 70)

        # Map of (type, instance) → list of package paths
        tgi_map: dict[tuple[int, int], list[Path]] = defaultdict(list)
        parsed = 0
        failed = 0

        for fp in tqdm(self.packages, desc="  Parsing package indexes"):
            parser = self._parse_package(fp)
            if parser is None:
                failed += 1
                continue
            parsed += 1

            for entry in parser.entries:
                # Only track tuning and simdata resources for conflicts
                if entry.res_type in (DBPFParser.TUNING_XML, DBPFParser.SIMDATA):
                    key = (entry.res_type, entry.instance)
                    tgi_map[key].append(fp)

        # Find conflicts (same resource in multiple packages)
        conflicts = {k: v for k, v in tgi_map.items() if len(v) > 1}

        print(f"\n  Parsed {parsed:,} packages ({failed} failed)")

        if conflicts:
            # Sort by number of conflicting files (most conflicts first)
            sorted_conflicts = sorted(conflicts.items(), key=lambda x: len(x[1]), reverse=True)
            print(f"\n  TUNING CONFLICTS — {len(conflicts)} resources overridden by multiple mods:")

            shown = 0
            for (res_type, instance), files in sorted_conflicts:
                if shown >= 50:
                    print(f"\n    ... and {len(sorted_conflicts) - 50} more conflicts")
                    break
                shown += 1
                type_name = DBPFParser.type_name(res_type)
                print(f"\n    {type_name} 0x{instance:016X} — {len(files)} mods conflict:")
                for fp in sorted(files):
                    rel = fp.relative_to(self.mods_path)
                    print(f"      [!] {rel}")
                    self._add_issue("tuning_conflict", fp, "Medium",
                                    "Tuning override conflict",
                                    f"{type_name} instance 0x{instance:016X} "
                                    f"overridden by {len(files)} packages",
                                    self._file_size(fp), self._file_date(fp))
                print(f"    → Keep only the mod you prefer, remove others")
        else:
            print("\n  No tuning conflicts detected.")

        return len(conflicts)

    # ── 7. LOD / Mesh Quality Analysis ──

    def analyze_mesh_quality(self) -> int:
        """Check CAS CC for missing LOD levels using actual GEOM resource counts.

        Properly-made CAS CC should have 3-4 LOD meshes per part. CC with only
        one GEOM but containing CAS parts likely has no LOD optimization."""
        print("\n" + "=" * 70)
        print("MESH / LOD QUALITY ANALYSIS")
        print("=" * 70)

        missing_lod = []
        oversized_mesh = []
        no_geom_cas = []

        cas_packages = 0
        for fp in tqdm(self.packages, desc="  Analyzing mesh data"):
            parser = self._parse_package(fp)
            if parser is None:
                continue

            type_counts = parser.get_resource_type_counts()
            type_sizes = parser.get_resource_type_sizes()

            cas_count = type_counts.get(DBPFParser.CAS_PART, 0)
            geom_count = type_counts.get(DBPFParser.GEOM, 0)
            geom_size = type_sizes.get(DBPFParser.GEOM, 0)

            if cas_count == 0:
                continue  # Not CAS CC, skip

            cas_packages += 1

            if geom_count == 0 and cas_count > 0:
                # CAS part with no geometry — might be an override or recolor
                # Only flag if the package is large (suggests missing content)
                if self._file_size(fp) > 50000:
                    no_geom_cas.append(fp)
                continue

            # Check LOD ratio: well-made CC typically has 3+ GEOMs per CAS part
            if cas_count > 0 and geom_count > 0:
                ratio = geom_count / cas_count
                if ratio < 2.0 and geom_size > 100000:
                    missing_lod.append((fp, cas_count, geom_count, geom_size))

            # Check for oversized meshes
            if geom_size > 5 * 1024 * 1024:  # 5MB+ of mesh data
                oversized_mesh.append((fp, geom_size))

        print(f"\n  Analyzed {cas_packages:,} CAS CC packages")

        if missing_lod:
            print(f"\n  LIKELY MISSING LODs — {len(missing_lod)} packages:")
            print("  (CAS CC with few GEOM resources relative to part count)")
            for fp, cas_n, geom_n, gsz in sorted(missing_lod, key=lambda x: x[3], reverse=True)[:30]:
                rel = fp.relative_to(self.mods_path)
                print(f"    [!] {rel}  ({cas_n} CAS parts, {geom_n} GEOMs, {gsz/1024:.0f} KB mesh)")
                self._add_issue("missing_lod", fp, "Low",
                                "Missing LOD levels (performance)",
                                f"{cas_n} CAS parts but only {geom_n} GEOMs ({gsz/1024:.0f} KB mesh)",
                                self._file_size(fp), self._file_date(fp))
            if len(missing_lod) > 30:
                print(f"    ... and {len(missing_lod) - 30} more")

        if oversized_mesh:
            print(f"\n  OVERSIZED MESHES — {len(oversized_mesh)} packages:")
            for fp, gsz in sorted(oversized_mesh, key=lambda x: x[1], reverse=True)[:20]:
                rel = fp.relative_to(self.mods_path)
                print(f"    [!] {rel}  ({gsz/1024/1024:.1f} MB of mesh data)")
                self._add_issue("oversized_mesh", fp, "Medium",
                                "Performance impact (large mesh)",
                                f"{gsz/1024/1024:.1f} MB of GEOM data",
                                self._file_size(fp), self._file_date(fp))

        total = len(missing_lod) + len(oversized_mesh)
        if total == 0:
            print("\n  All CAS CC mesh quality looks acceptable.")
        return total

    # ── 8. Performance Profiling ──

    def profile_performance(self) -> int:
        """Identify performance-heavy packages by resource composition."""
        print("\n" + "=" * 70)
        print("PERFORMANCE PROFILING")
        print("=" * 70)

        fat_files = []       # Files over 5 MB
        extreme_files = []   # Files over 50 MB
        high_resource = []   # Files with excessive resource counts
        script_count = len(self.scripts)

        for fp in self.packages:
            size = self._file_size(fp)
            size_mb = size / (1024 * 1024)

            if size_mb > 50:
                extreme_files.append((fp, size_mb))
                self._add_issue("performance", fp, "High",
                                "Severe load time / memory impact",
                                f"Extremely large: {size_mb:.1f} MB",
                                size, self._file_date(fp))
            elif size_mb > 5:
                fat_files.append((fp, size_mb))
                self._add_issue("performance", fp, "Low",
                                "Load time impact",
                                f"Large file: {size_mb:.1f} MB",
                                size, self._file_date(fp))

            # Check resource count (only on already-parsed packages)
            if fp in self.dbpf_cache:
                parser = self.dbpf_cache[fp]
                if len(parser.entries) > 5000:
                    high_resource.append((fp, len(parser.entries)))

        if extreme_files:
            extreme_files.sort(key=lambda x: x[1], reverse=True)
            print(f"\n  EXTREMELY LARGE (>50 MB) — {len(extreme_files)} files:")
            for fp, sz in extreme_files:
                rel = fp.relative_to(self.mods_path)
                print(f"    [X] {rel} — {sz:.1f} MB")

        if fat_files:
            fat_files.sort(key=lambda x: x[1], reverse=True)
            total_fat = sum(sz for _, sz in fat_files)
            print(f"\n  LARGE FILES (5-50 MB) — {len(fat_files)} files, {total_fat:.0f} MB total:")
            for fp, sz in fat_files[:20]:
                rel = fp.relative_to(self.mods_path)
                print(f"    [!] {rel} — {sz:.1f} MB")
            if len(fat_files) > 20:
                print(f"    ... and {len(fat_files) - 20} more")

        if high_resource:
            print(f"\n  HIGH RESOURCE COUNT — {len(high_resource)} packages:")
            for fp, count in sorted(high_resource, key=lambda x: x[1], reverse=True)[:10]:
                rel = fp.relative_to(self.mods_path)
                print(f"    [!] {rel} — {count:,} resources")

        if script_count > 100:
            print(f"\n  [!] HIGH SCRIPT COUNT: {script_count} .ts4script files")
            print("      This may cause simulation lag. Consider removing unused script mods.")
        elif script_count > 50:
            print(f"\n  [i] Script count: {script_count} (moderate — monitor for lag)")
        else:
            print(f"\n  Script count: {script_count} (OK)")

        total = len(extreme_files) + len(fat_files) + len(high_resource)
        if total == 0 and script_count <= 50:
            print("\n  Performance profile looks good.")
        return total

    # ── 9. Exception Log Analysis ──

    def analyze_exception_log(self) -> int:
        """Parse Better Exceptions / MCCC last exception logs for useful info."""
        if not self.log_path or not self.log_path.exists():
            print("\n" + "=" * 70)
            print("EXCEPTION LOG ANALYSIS — Skipped (no log file provided)")
            print("=" * 70)
            return 0

        print("\n" + "=" * 70)
        print("EXCEPTION LOG ANALYSIS")
        print("=" * 70)

        try:
            with open(self.log_path, "r", encoding="utf-8", errors="ignore") as f:
                content = f.read()
        except Exception as e:
            print(f"  Error reading log: {e}")
            return 0

        issues_found = 0

        # Better Exceptions advice
        be_advice = re.search(r"<Advice>(.*?)</Advice>", content, re.DOTALL)
        if be_advice:
            advice = be_advice.group(1).strip()
            print(f"\n  Better Exceptions diagnosis:")
            print(f"    {advice[:200]}")
            self._add_issue("log_analysis", self.log_path, "Info", "Diagnostic",
                            f"BetterExceptions: {advice[:200]}")
            issues_found += 1

        # Bad CC flag
        bad_cc = re.search(r"<BadObjectCC>(True)</BadObjectCC>", content)
        if bad_cc:
            print(f"\n  [!] BAD OBJECT CC DETECTED in crash log")
            print(f"      A broken custom content object caused this crash.")
            self._add_issue("log_analysis", self.log_path, "High",
                            "Object crash", "BadObjectCC flag set in crash log")
            issues_found += 1

        # Error category
        category = re.search(r"<categoryid>(.*?)</categoryid>", content)
        if category:
            loc = category.group(1).strip()
            print(f"\n  Error location: {loc}")
            self._add_issue("log_analysis", self.log_path, "High", "Script error",
                            f"Error in: {loc}")
            issues_found += 1

        # Crash error message
        crash_msg = re.search(r"\[BE Interceptor\].*?\((.*?)\)", content)
        if crash_msg:
            err = crash_msg.group(1).strip()
            print(f"\n  Crash error: {err[:150]}")
            self._add_issue("log_analysis", self.log_path, "High", "Crash",
                            f"Crash: {err[:150]}")
            issues_found += 1

            # Specific: NoneType outfit crash
            if "'NoneType' object has no attribute 'back'" in err:
                print(f"\n  → OUTFIT/CAS ERROR: A Sim has broken clothing/hair/accessories")
                print(f"    Fix: Remove CAS CC older than 2 years, especially hair and clothing")

        # Traceback — extract ACTUAL MOD names from paths, not EA game scripts
        mod_mentions = set()

        # EA/game script paths to ignore — these are the game engine, not mods
        EA_SCRIPT_DIRS = {
            "server", "core", "t:\\ingame", "gameplay\\scripts",
        }
        EA_SCRIPT_FILES = {
            "alarms", "areaserver", "base_statistic_listener", "buff",
            "buff_component", "commodity", "continuous_statistic",
            "injector", "loot", "loot_basic_op", "resolver",
            "scheduling", "tests", "time_service", "utils", "wrapper",
            "zone", "zone_tick", "__init__", "exported_func",
        }

        # Windows mod paths: look for files under a Mods\ directory
        for match in re.finditer(r"[A-Z]:\\[^\\]+\\[^\\]+\\Mods\\([^\\]+)", content):
            mod_mentions.add(match.group(1))

        # Python files in traceback — only flag if NOT an EA game script
        for match in re.finditer(
                r'File\s+"([^"]*?[/\\]([^/\\"]+?)\.py)"', content):
            full_path = match.group(1).lower().replace("\\", "/")
            basename = match.group(2)

            # Skip EA game engine scripts
            if basename.lower() in EA_SCRIPT_FILES:
                continue
            if any(d in full_path for d in EA_SCRIPT_DIRS):
                continue

            # This is likely an actual mod script
            mod_mentions.add(basename)

        if mod_mentions:
            print(f"\n  Mods referenced in crash trace ({len(mod_mentions)}):")
            for mod in sorted(mod_mentions):
                print(f"    [!] {mod}")
                self._add_issue("log_analysis", self.log_path, "High",
                                "In crash trace", f"Mod in traceback: {mod}")
            issues_found += len(mod_mentions)

        if issues_found == 0:
            print("\n  No actionable issues found in log.")
        return issues_found

    # ── 10. Resource Composition Analysis ──

    def analyze_resource_composition(self) -> None:
        """Print a summary of resource types across all parsed packages.
        Informational only — does not flag issues."""
        print("\n" + "=" * 70)
        print("RESOURCE COMPOSITION SUMMARY")
        print("=" * 70)

        total_counts = defaultdict(int)
        total_sizes = defaultdict(int)

        for parser in self.dbpf_cache.values():
            for entry in parser.entries:
                total_counts[entry.res_type] += 1
                total_sizes[entry.res_type] += entry.file_size

        if not total_counts:
            print("\n  No packages were parsed (run conflict detection first).")
            return

        print(f"\n  {'Resource Type':<25} {'Count':>10} {'Total Size':>12}")
        print(f"  {'-'*25} {'-'*10} {'-'*12}")

        for type_id, count in sorted(total_counts.items(), key=lambda x: x[1], reverse=True)[:20]:
            name = DBPFParser.type_name(type_id)
            size = total_sizes[type_id]
            print(f"  {name:<25} {count:>10,} {size/1024/1024:>10.1f} MB")

        total_resources = sum(total_counts.values())
        total_size = sum(total_sizes.values())
        print(f"\n  Total: {total_resources:,} resources, {total_size/1024/1024:.1f} MB indexed data")

    # ── 11. Script Mod Version Detection ──

    def detect_mod_versions(self) -> int:
        """Read actual version strings from script mods and tuning XML to
        determine if known script mods are up-to-date.

        Checks:
        1. Filename version patterns (e.g. UI_Cheats_Extension_v1.43.package)
        2. Tuning XML content inside .package files for version strings
        3. Python source inside .ts4script (ZIP) files for version variables
        """
        print("\n" + "=" * 70)
        print("SCRIPT MOD VERSION DETECTION")
        print("=" * 70)

        all_files = self.packages + self.scripts
        detections = []  # (file, mod_info, version_tuple, version_str, source)

        for fp in all_files:
            fn_lower = fp.stem.lower().replace(" ", "").replace("-", "").replace("_", "")
            # Also get full relative path for checking parent folders
            full_path_lower = str(fp).lower()

            for mod_info in KNOWN_SCRIPT_MODS:
                # Check if this file matches any known script mod
                matched = False
                for match_str in mod_info["filename_match"]:
                    clean_match = match_str.replace("_", "").replace("-", "")
                    if clean_match in fn_lower:
                        matched = True
                        break
                if not matched:
                    continue

                # Check exclude patterns (e.g. WW animation packs)
                excludes = mod_info.get("exclude_substrings", [])
                if excludes:
                    excluded = False
                    for excl in excludes:
                        if excl in full_path_lower:
                            excluded = True
                            break
                    if excluded:
                        continue

                # Try to extract version
                version_tuple = None
                version_str = ""
                source = ""

                # Method 1: Check filename for version pattern
                for pattern in mod_info["version_patterns"]:
                    m = re.search(pattern, fp.stem, re.IGNORECASE)
                    if m:
                        groups = [int(g) for g in m.groups() if g is not None]
                        while len(groups) < 3:
                            groups.append(0)
                        version_tuple = tuple(groups[:3])
                        version_str = ".".join(str(g) for g in version_tuple)
                        source = "filename"
                        break

                # Method 1b: Check parent folder name for version pattern
                # (e.g. McCmdCenter_AllModules_2025_7_0\mc_cmd_center.package)
                if version_tuple is None:
                    for parent in fp.parents:
                        if parent == self.mods_path:
                            break
                        folder_name = parent.name
                        for pattern in mod_info["version_patterns"]:
                            m = re.search(pattern, folder_name, re.IGNORECASE)
                            if m:
                                groups = [int(g) for g in m.groups()
                                          if g is not None]
                                while len(groups) < 3:
                                    groups.append(0)
                                version_tuple = tuple(groups[:3])
                                version_str = ".".join(
                                    str(g) for g in version_tuple)
                                source = f"folder:{folder_name}"
                                break
                        if version_tuple:
                            break

                # Method 2: For .ts4script files, peek inside the ZIP
                if version_tuple is None and fp.suffix.lower() == ".ts4script":
                    version_tuple, version_str, source = self._extract_script_version(
                        fp, mod_info)

                # Method 3: For .package files, read tuning XML resources
                if version_tuple is None and fp.suffix.lower() == ".package":
                    version_tuple, version_str, source = self._extract_package_version(
                        fp, mod_info)

                detections.append((fp, mod_info, version_tuple, version_str, source))
                break  # Only match one mod per file

        if not detections:
            print("\n  No known script mods found to version-check.")
            return 0

        # Group by mod name
        by_mod = defaultdict(list)
        for fp, mod_info, vtup, vstr, src in detections:
            by_mod[mod_info["name"]].append((fp, mod_info, vtup, vstr, src))

        outdated_count = 0
        print(f"\n  Found {len(detections)} known script mod files across "
              f"{len(by_mod)} mods:\n")

        for mod_name, entries in sorted(by_mod.items()):
            mod_info = entries[0][1]
            min_safe = mod_info["min_safe_version"]
            min_safe_str = ".".join(str(v) for v in min_safe)

            print(f"  {mod_name} (by {mod_info['creator']})")
            print(f"    Minimum safe version: {min_safe_str}")

            for fp, _, vtup, vstr, src in entries:
                rel = fp.relative_to(self.mods_path)
                if vtup:
                    is_safe = vtup >= min_safe
                    status = "OK" if is_safe else "OUTDATED"
                    icon = "  " if is_safe else "[!]"
                    print(f"    {icon} {rel}")
                    print(f"        Version: {vstr} (from {src}) — {status}")

                    if not is_safe:
                        outdated_count += 1
                        self._add_issue("version_outdated", fp, "Medium",
                                        mod_info.get("notes", "Needs update"),
                                        f"{mod_name} v{vstr} is below minimum "
                                        f"safe v{min_safe_str} | {mod_info['notes']}",
                                        self._file_size(fp), self._file_date(fp))
                    else:
                        self._add_issue("version_ok", fp, "Info",
                                        "Version confirmed current",
                                        f"{mod_name} v{vstr} — up to date "
                                        f"(min: v{min_safe_str})",
                                        self._file_size(fp), self._file_date(fp))
                else:
                    print(f"    [?] {rel}")
                    print(f"        Version: could not detect")
                    self._add_issue("version_unknown", fp, "Low",
                                    "Version unreadable",
                                    f"{mod_name} — version not detected in filename "
                                    f"or metadata. Manually verify ≥ v{min_safe_str}",
                                    self._file_size(fp), self._file_date(fp))
            print()

        if outdated_count:
            print(f"  {outdated_count} script mod(s) appear OUTDATED.")
        else:
            print(f"  All detected script mod versions look current.")
        return outdated_count

    def _extract_script_version(self, filepath: Path,
                                mod_info: dict) -> tuple:
        """Open a .ts4script (ZIP) and search Python source for version strings."""
        try:
            with zipfile.ZipFile(filepath, "r") as zf:
                for name in zf.namelist():
                    if not name.endswith(".py"):
                        continue
                    try:
                        src = zf.read(name).decode("utf-8", errors="ignore")
                    except Exception:
                        continue

                    # Common version variable patterns in Python source
                    for pattern in [
                        r"__version__\s*=\s*['\"]([^'\"]+)['\"]",
                        r"VERSION\s*=\s*['\"]([^'\"]+)['\"]",
                        r"version\s*=\s*['\"](\d+[.\d]+)['\"]",
                        r"MOD_VERSION\s*=\s*['\"]([^'\"]+)['\"]",
                        r"mod_version\s*=\s*['\"]([^'\"]+)['\"]",
                        r"CURRENT_VERSION\s*=\s*['\"]([^'\"]+)['\"]",
                    ]:
                        m = re.search(pattern, src)
                        if m:
                            raw = m.group(1).strip()
                            nums = re.findall(r"\d+", raw)
                            if nums:
                                parts = [int(n) for n in nums[:3]]
                                while len(parts) < 3:
                                    parts.append(0)
                                return (tuple(parts[:3]),
                                        ".".join(str(p) for p in parts[:3]),
                                        f"script:{name}")

                    # Also try the mod's own patterns against the source
                    for pattern in mod_info.get("version_patterns", []):
                        m = re.search(pattern, src, re.IGNORECASE)
                        if m:
                            groups = [int(g) for g in m.groups() if g is not None]
                            while len(groups) < 3:
                                groups.append(0)
                            return (tuple(groups[:3]),
                                    ".".join(str(g) for g in groups[:3]),
                                    f"script:{name}")

        except (zipfile.BadZipFile, Exception):
            pass
        return (None, "", "")

    def _extract_package_version(self, filepath: Path,
                                 mod_info: dict) -> tuple:
        """Read tuning XML resources from a parsed package looking for
        version strings embedded in the XML content.

        IMPORTANT: Sims 4 tuning XML has a root-level 'version' attribute
        that indicates the tuning FORMAT version (always small numbers like
        1, 2, 3). This is NOT the mod version. We skip these by:
        1. Ignoring version= in the first XML tag (root element)
        2. Ignoring version values that are just a single digit
        3. Prioritizing explicit mod_version / MOD_VERSION / __version__ patterns
        """
        parser = self._parse_package(filepath)
        if parser is None:
            return (None, "", "")

        # Only check tuning/simdata resources
        tuning_entries = parser.get_by_type(DBPFParser.TUNING_XML)
        # Limit to first 20 tuning entries for speed
        for entry in tuning_entries[:20]:
            data = parser.read_resource(entry)
            if data is None:
                continue
            try:
                text = data.decode("utf-8", errors="ignore")
            except Exception:
                continue

            # Search for EXPLICIT mod version markers (high confidence)
            for pattern in [
                r"mod_version\s*=\s*['\"](\d+[.\d]+)['\"]",
                r"MOD_VERSION\s*=\s*['\"](\d+[.\d]+)['\"]",
                r"<T\s+n=\"mod_version\">([^<]+)</T>",
                r"<T\s+n=\"version\">([^<]+)</T>",
            ]:
                m = re.search(pattern, text, re.IGNORECASE)
                if m:
                    raw = m.group(1).strip()
                    nums = re.findall(r"\d+", raw)
                    if nums and not (len(nums) == 1 and int(nums[0]) <= 5):
                        # Skip single-digit values (likely format version)
                        parts = [int(n) for n in nums[:3]]
                        while len(parts) < 3:
                            parts.append(0)
                        return (tuple(parts[:3]),
                                ".".join(str(p) for p in parts[:3]),
                                f"tuning:0x{entry.instance:016X}")

            # Also try mod-specific patterns against content
            # Strip ALL version="..." attributes from XML — these are ALWAYS
            # tuning format versions (1, 2, 3), never mod versions.
            # They appear on multiple elements, not just the root.
            stripped = re.sub(r'\s+version\s*=\s*["\'][^"\']*["\']', '', text)
            for pattern in mod_info.get("version_patterns", []):
                m = re.search(pattern, stripped, re.IGNORECASE)
                if m:
                    groups = [int(g) for g in m.groups() if g is not None]
                    if groups and not (len(groups) == 1 and groups[0] <= 5):
                        while len(groups) < 3:
                            groups.append(0)
                        return (tuple(groups[:3]),
                                ".".join(str(g) for g in groups[:3]),
                                f"tuning:0x{entry.instance:016X}")

        return (None, "", "")

    # ── 12. Merged Package Analyzer ──

    def analyze_merged_packages(self, threshold_mb: float = 100.0) -> int:
        """Deep analysis of large/merged packages: break down resource
        composition so you can see what's eating space inside those giant files.

        For each package above threshold_mb, reports:
        - Total resource count and types
        - CAS part count (how many pieces of CC are merged in)
        - Mesh data size (GEOM)
        - Texture/image data size (DDS + PNG)
        - Thumbnail bloat (small/medium/large thumbnails)
        - Tuning/SimData (mod logic)
        - Animation clips
        - Estimated quality score and optimization suggestions
        """
        print("\n" + "=" * 70)
        print(f"MERGED PACKAGE ANALYSIS (files > {threshold_mb:.0f} MB)")
        print("=" * 70)

        large_packages = []
        for fp in self.packages:
            sz = self._file_size(fp)
            if sz / (1024 * 1024) >= threshold_mb:
                large_packages.append((fp, sz))

        if not large_packages:
            print(f"\n  No packages above {threshold_mb:.0f} MB threshold.")
            return 0

        large_packages.sort(key=lambda x: x[1], reverse=True)
        total_large_size = sum(sz for _, sz in large_packages)

        print(f"\n  Found {len(large_packages)} packages above {threshold_mb:.0f} MB")
        print(f"  Combined size: {total_large_size/1024**3:.1f} GB")
        print(f"  ({100*total_large_size/self.total_size_bytes:.0f}% of total mods folder)")

        results = []

        for fp, file_size in tqdm(large_packages, desc="  Analyzing merged packages"):
            parser = self._parse_package(fp)
            if parser is None:
                continue

            # Categorize every resource
            cats = {
                "CAS Parts": (DBPFParser.CAS_PART, 0, 0),
                "Meshes (GEOM)": (DBPFParser.GEOM, 0, 0),
                "DDS Textures": (DBPFParser.IMG_DDS, 0, 0),
                "PNG Images": (DBPFParser.IMG_PNG, 0, 0),
                "Thumb Small": (DBPFParser.THUMB_SM, 0, 0),
                "Thumb Medium": (DBPFParser.THUMB_MD, 0, 0),
                "Thumb Large": (DBPFParser.THUMB_LG, 0, 0),
                "Tuning XML": (DBPFParser.TUNING_XML, 0, 0),
                "SimData": (DBPFParser.SIMDATA, 0, 0),
                "Animations": (DBPFParser.CLIP, 0, 0),
                "Models": (DBPFParser.MODL, 0, 0),
                "String Tables": (DBPFParser.STBL, 0, 0),
                "Object Defs": (DBPFParser.OBJDEF, 0, 0),
            }

            # Count and sum sizes per category
            breakdown = {}
            other_count = 0
            other_size = 0
            for cat_name, (type_id, _, _) in cats.items():
                count = 0
                size = 0
                for e in parser.entries:
                    if e.res_type == type_id:
                        count += 1
                        size += e.file_size
                breakdown[cat_name] = (count, size)

            # Count "other"
            known_types = {v[0] for v in cats.values()}
            for e in parser.entries:
                if e.res_type not in known_types:
                    other_count += 1
                    other_size += e.file_size
            breakdown["Other"] = (other_count, other_size)

            # Calculate thumbnail bloat
            thumb_size = sum(breakdown[k][1] for k in
                             ["Thumb Small", "Thumb Medium", "Thumb Large"])
            thumb_pct = (100 * thumb_size / file_size) if file_size > 0 else 0

            # Calculate mesh weight
            mesh_size = breakdown["Meshes (GEOM)"][1]
            mesh_pct = (100 * mesh_size / file_size) if file_size > 0 else 0

            # Texture weight
            tex_size = breakdown["DDS Textures"][1] + breakdown["PNG Images"][1]
            tex_pct = (100 * tex_size / file_size) if file_size > 0 else 0

            cas_count = breakdown["CAS Parts"][0]
            total_resources = len(parser.entries)

            # Optimization suggestions
            suggestions = []
            if thumb_pct > 30:
                suggestions.append(f"Thumbnail bloat ({thumb_pct:.0f}%) — "
                                   f"re-merge with Sims 4 Studio thumbnail removal")
            if mesh_pct > 50:
                suggestions.append(f"Mesh-heavy ({mesh_pct:.0f}%) — "
                                   f"check for high-poly CC without LODs")
            if tex_pct > 60:
                suggestions.append(f"Texture-heavy ({tex_pct:.0f}%) — "
                                   f"some textures may be higher res than needed")
            if file_size > 1.5 * 1024**3:  # 1.5 GB
                suggestions.append(f"Near DBPF 2GB limit — consider splitting")
            if cas_count > 3000:
                suggestions.append(f"{cas_count:,} CAS parts — very large merge, "
                                   f"will slow CAS loading")
            if not suggestions:
                suggestions.append("Within normal range for merged CC")

            results.append({
                "path": fp,
                "file_size": file_size,
                "total_resources": total_resources,
                "cas_count": cas_count,
                "breakdown": breakdown,
                "thumb_size": thumb_size,
                "thumb_pct": thumb_pct,
                "mesh_size": mesh_size,
                "mesh_pct": mesh_pct,
                "tex_size": tex_size,
                "tex_pct": tex_pct,
                "suggestions": suggestions,
            })

        # Print results
        print(f"\n{'=' * 70}")
        for r in results:
            rel = r["path"].relative_to(self.mods_path)
            sz_mb = r["file_size"] / (1024**2)
            sz_gb = r["file_size"] / (1024**3)

            if sz_gb >= 1.0:
                sz_label = f"{sz_gb:.1f} GB"
            else:
                sz_label = f"{sz_mb:.0f} MB"

            print(f"\n  {rel}")
            print(f"  {'─' * 60}")
            print(f"  Total: {sz_label} | "
                  f"{r['total_resources']:,} resources | "
                  f"{r['cas_count']:,} CAS parts")
            print()

            # Resource breakdown table
            print(f"    {'Category':<20} {'Count':>8} {'Size':>10} {'% of File':>10}")
            print(f"    {'─'*20} {'─'*8} {'─'*10} {'─'*10}")

            for cat_name, (count, size) in sorted(
                    r["breakdown"].items(),
                    key=lambda x: x[1][1], reverse=True):
                if count == 0:
                    continue
                pct = (100 * size / r["file_size"]) if r["file_size"] > 0 else 0
                if size > 1024 * 1024:
                    size_str = f"{size/1024/1024:.1f} MB"
                elif size > 1024:
                    size_str = f"{size/1024:.0f} KB"
                else:
                    size_str = f"{size} B"
                print(f"    {cat_name:<20} {count:>8,} {size_str:>10} {pct:>9.1f}%")

            # Summary bars
            print()
            print(f"    Thumbnails: {r['thumb_size']/1024/1024:.0f} MB "
                  f"({r['thumb_pct']:.0f}%)", end="")
            if r["thumb_pct"] > 30:
                print("  ← BLOATED", end="")
            print()
            print(f"    Meshes:     {r['mesh_size']/1024/1024:.0f} MB "
                  f"({r['mesh_pct']:.0f}%)", end="")
            if r["mesh_pct"] > 50:
                print("  ← HEAVY", end="")
            print()
            print(f"    Textures:   {r['tex_size']/1024/1024:.0f} MB "
                  f"({r['tex_pct']:.0f}%)", end="")
            if r["tex_pct"] > 60:
                print("  ← HEAVY", end="")
            print()

            # Suggestions
            print()
            for s in r["suggestions"]:
                print(f"    → {s}")

            # Add CSV issues for notable findings
            for s in r["suggestions"]:
                if any(w in s.lower() for w in ["bloat", "limit", "heavy", "slow"]):
                    self._add_issue("merged_analysis", r["path"], "Medium",
                                    "Optimization opportunity",
                                    f"{sz_label}, {r['cas_count']:,} CAS parts | {s}",
                                    r["file_size"], self._file_date(r["path"]))
                    break  # Only one issue per file

        # Summary
        print(f"\n{'=' * 70}")
        print(f"  MERGED PACKAGE SUMMARY")
        print(f"  {'─' * 50}")
        total_cas = sum(r["cas_count"] for r in results)
        total_thumb = sum(r["thumb_size"] for r in results)
        total_mesh = sum(r["mesh_size"] for r in results)
        total_tex = sum(r["tex_size"] for r in results)
        print(f"  Packages analyzed:     {len(results)}")
        print(f"  Combined size:         {total_large_size/1024**3:.1f} GB")
        print(f"  Total CAS parts:       {total_cas:,}")
        print(f"  Total thumbnail data:  {total_thumb/1024**3:.1f} GB")
        print(f"  Total mesh data:       {total_mesh/1024**3:.1f} GB")
        print(f"  Total texture data:    {total_tex/1024**3:.1f} GB")

        if total_thumb / (1024**3) > 2.0:
            print(f"\n  [!] Your merged packages contain {total_thumb/1024**3:.1f} GB "
                  f"of thumbnails alone.")
            print(f"      Re-merging with thumbnail removal could save significant space.")

        return len(results)

    # ── 13. Full Report ──

    def generate_report(self):
        """Print final summary with health score."""
        print("\n" + "=" * 70)
        print("FINAL REPORT")
        print("=" * 70)

        high = sum(1 for i in self.issues if i["Risk"] == "High")
        medium = sum(1 for i in self.issues if i["Risk"] == "Medium")
        low = sum(1 for i in self.issues if i["Risk"] == "Low")
        info = sum(1 for i in self.issues if i["Risk"] == "Info")

        # Health score: start at 100, deduct by severity
        score = max(0, 100 - (high * 5) - (medium * 2) - (low * 1))
        score = min(100, score)

        print(f"\n  Packages scanned:  {len(self.packages):,}")
        print(f"  Scripts scanned:   {len(self.scripts):,}")
        print(f"  Total mod size:    {self.total_size_mb:.1f} MB ({self.total_size_mb/1024:.2f} GB)")
        print(f"  DBPF parse errors: {len(self.parse_failures)}")
        print()
        print(f"  Issues found:")
        print(f"    High risk:   {high}")
        print(f"    Medium risk: {medium}")
        print(f"    Low risk:    {low}")
        print(f"    Info:        {info}")
        print()
        print(f"  HEALTH SCORE: {score}/100", end="")
        if score >= 80:
            print("  — Good")
        elif score >= 50:
            print("  — Needs attention")
        else:
            print("  — Critical cleanup needed")

        print(f"\n  PRIORITY ACTIONS:")
        print(f"  {'─'*50}")
        if high > 0:
            print(f"  1. Fix {high} high-risk issues (broken mods, integrity failures)")
        if medium > 0:
            print(f"  2. Address {medium} medium-risk issues (duplicates, outdated mods)")
        if low > 0:
            print(f"  3. Consider {low} low-risk optimizations (LODs, large files)")

        print(f"\n  BACKUP REMINDER:")
        print(f"  {'─'*50}")
        print(f"  Before making changes: copy your entire Mods folder to Mods_Backup")
        print(f"  Test the game after each batch of removals")

    # ── 12. CSV Export ──

    def export_to_csv(self, filename: str):
        """Export all issues with metadata to CSV."""
        if not self.issues:
            print(f"  No issues to export.")
            return

        with open(filename, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=["Type", "File", "Size",
                                                     "Date", "Notes", "Risk", "Symptom"])
            writer.writeheader()
            for row in self.issues:
                writer.writerow(row)

        print(f"  Exported {len(self.issues)} issues to {filename}")

    def export_cleanup_xlsx(self, csv_path: str):
        """Generate a formatted Excel cleanup checklist from the CSV report."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("  [Skip] openpyxl not installed — run 'pip install openpyxl' for Excel cleanup lists")
            return

        # Read the CSV we just exported
        rows_by_type = {}
        try:
            with open(csv_path, "r", encoding="utf-8-sig") as f:
                for row in csv.DictReader(f):
                    t = row.get("Type", "").strip()
                    rows_by_type.setdefault(t, []).append(row)
        except Exception as e:
            print(f"  [Error] Could not read CSV for Excel export: {e}")
            return

        # Styles
        RED_FILL = PatternFill('solid', fgColor='FFCCCC')
        ORANGE_FILL = PatternFill('solid', fgColor='FFE0B2')
        YELLOW_FILL = PatternFill('solid', fgColor='FFF9C4')
        BLUE_FILL = PatternFill('solid', fgColor='BBDEFB')
        GREEN_FILL = PatternFill('solid', fgColor='C8E6C9')
        HEADER_FILL = PatternFill('solid', fgColor='263238')
        ALT_ROW = PatternFill('solid', fgColor='F5F5F5')
        WHITE_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        TITLE_FONT = Font(name='Arial', bold=True, size=14, color='263238')
        BOLD_FONT = Font(name='Arial', bold=True, size=10)
        NORMAL_FONT = Font(name='Arial', size=10)
        SMALL_FONT = Font(name='Arial', size=9, color='666666')
        THIN = Border(
            bottom=Side(style='thin', color='D0D0D0'),
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
        )
        RISK_FILLS = {'CRITICAL': RED_FILL, 'HIGH': ORANGE_FILL, 'MEDIUM': YELLOW_FILL,
                      'LOW': ALT_ROW, 'OK': GREEN_FILL}

        def _basename(path):
            return path.replace('/', '\\').split('\\')[-1]

        def _subfolder(path):
            if '\\Mods\\' in path:
                after = path.split('\\Mods\\', 1)[-1]
                parts = after.replace('/', '\\').split('\\')
                return parts[0] if len(parts) > 1 else ''
            return ''

        def _header_row(ws, row, headers):
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=ci, value=h)
                c.fill = HEADER_FILL
                c.font = WHITE_FONT
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                c.border = THIN

        def _data_row(ws, row, values, fill=None, wrap=False):
            for ci, v in enumerate(values, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.font = NORMAL_FONT
                c.border = THIN
                c.alignment = Alignment(vertical='center', wrap_text=wrap)
                if fill:
                    c.fill = fill

        def _set_widths(ws, widths):
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        wb = Workbook()
        today = datetime.now().strftime("%B %d, %Y")

        # ── SHEET 1: SUMMARY ──
        ws = wb.active
        ws.title = "Summary"
        ws.sheet_properties.tabColor = "4CAF50"
        ws['A1'] = "MOD HEALTH SUMMARY"
        ws['A1'].font = TITLE_FONT
        ws.row_dimensions[1].height = 30
        ws['A2'] = f"Generated: {today} | Scanner v{VERSION}"
        ws['A2'].font = SMALL_FONT

        stats = [
            ('Crash errors in logs', len(rows_by_type.get('log_analysis', [])), 'CRITICAL'),
            ('Critically outdated CC (10+ yrs)', len(rows_by_type.get('outdated_critical', [])), 'HIGH'),
            ('Known broken mods', len(rows_by_type.get('known_broken', [])), 'HIGH'),
            ('Mod Checker — Broken', len(rows_by_type.get('modlist_broken', [])), 'HIGH'),
            ('Mod Checker — Obsolete', len(rows_by_type.get('modlist_obsolete', [])), 'HIGH'),
            ('Mod Checker — Minor Issues', len(rows_by_type.get('modlist_minor_issue', [])), 'MEDIUM'),
            ('Mod Checker — Unknown Status', len(rows_by_type.get('modlist_unknown', [])), 'LOW'),
            ('Script mods — version unknown', len(rows_by_type.get('version_unknown', [])), 'MEDIUM'),
            ('Script mods — version OK', len(rows_by_type.get('version_ok', [])), 'OK'),
            ('Script mods — outdated', len(rows_by_type.get('version_outdated', [])), 'MEDIUM'),
            ('Outdated CC (medium risk)', len(rows_by_type.get('outdated', [])), 'MEDIUM'),
            ('Tuning conflicts', len(rows_by_type.get('tuning_conflict', [])), 'MEDIUM'),
            ('Performance warnings', len(rows_by_type.get('performance', [])), 'LOW'),
            ('Missing LODs', len(rows_by_type.get('missing_lod', [])), 'LOW'),
            ('Oversized meshes', len(rows_by_type.get('oversized_mesh', [])), 'LOW'),
            ('Oversized files (100MB+)', len(rows_by_type.get('merged_analysis', [])), 'MEDIUM'),
        ]

        r = 4
        _header_row(ws, r, ['Category', 'Count', 'Risk Level'])
        r += 1
        for label, count, risk in stats:
            _data_row(ws, r, [label, count, risk], RISK_FILLS.get(risk))
            r += 1
        r += 1
        total = sum(len(v) for v in rows_by_type.values())
        _data_row(ws, r, ['TOTAL FILES SCANNED', total, ''])
        for ci in range(1, 4):
            ws.cell(row=r, column=ci).font = BOLD_FONT
        _set_widths(ws, [40, 10, 14])
        ws.freeze_panes = 'A5'

        # ── SHEET 2: ACTION CHECKLIST ──
        ws2 = wb.create_sheet("Action Checklist")
        ws2.sheet_properties.tabColor = "FF5722"
        ws2['A1'] = "SIMS 4 MOD CLEANUP CHECKLIST"
        ws2['A1'].font = TITLE_FONT
        ws2.row_dimensions[1].height = 30
        ws2['A2'] = "Work through sections in order. Check off items as you complete them."
        ws2['A2'].font = SMALL_FONT

        r = 4
        hdrs = ['\u2713', 'Action', 'Mod Name', 'Problem', 'Risk', 'Fix']

        # --- P1: CRASH & BROKEN ---
        crash_items = []
        for item in rows_by_type.get('log_analysis', []):
            notes = item.get('Notes', '')
            if 'Crash:' in notes:
                crash_items.append(['\u2610', 'DIAGNOSE', 'Crash Error',
                                    notes.replace('Crash: ', '')[:80], 'HIGH', 'See traceback mods below'])
            elif 'Mod in traceback:' in notes:
                mod = notes.replace('Mod in traceback: ', '')
                crash_items.append(['\u2610', 'CHECK', f'Traceback: {mod}',
                                    'In crash call chain', 'HIGH', 'Update or verify this mod'])
        for item in rows_by_type.get('known_broken', []):
            notes = item.get('Notes', '')
            mod_name = notes.split(' \u2014 ')[0] if ' \u2014 ' in notes else _basename(item.get('File', ''))
            fix = notes.split('Fix: ')[-1] if 'Fix: ' in notes else 'Update to latest'
            crash_items.append(['\u2610', 'UPDATE', mod_name,
                                notes.split(' | ')[0][:80] if ' | ' in notes else notes[:80],
                                item.get('Risk', 'Medium'), fix])
        for item in rows_by_type.get('version_outdated', []):
            crash_items.append(['\u2610', 'UPDATE', _basename(item.get('File', '')),
                                item.get('Notes', '')[:80], 'MEDIUM', 'Download latest version'])
        # Add mods flagged Broken/Obsolete by Mod Checker List
        for item in rows_by_type.get('modlist_broken', []):
            notes = item.get('Notes', '')
            fp = item.get('File', '')
            # Notes format: "filename → Mod Name by Creator — BROKEN | details"
            your_file = _basename(fp)
            # Extract mod name and creator from notes
            if ' → ' in notes:
                mod_info = notes.split(' → ')[1].split(' — ')[0] if ' — ' in notes else notes.split(' → ')[1][:60]
            else:
                mod_info = notes.split(' — ')[0] if ' — ' in notes else your_file
            detail = notes.split('|')[-1].strip()[:80] if '|' in notes else 'Check mod page'
            crash_items.append(['\u2610', 'REMOVE/UPDATE', your_file,
                                f'BROKEN per Mod Checker → {mod_info}',
                                'HIGH', detail])
        for item in rows_by_type.get('modlist_obsolete', []):
            notes = item.get('Notes', '')
            fp = item.get('File', '')
            your_file = _basename(fp)
            if ' → ' in notes:
                mod_info = notes.split(' → ')[1].split(' — ')[0] if ' — ' in notes else notes.split(' → ')[1][:60]
            else:
                mod_info = notes.split(' — ')[0] if ' — ' in notes else your_file
            detail = notes.split('|')[-1].strip()[:80] if '|' in notes else 'No longer maintained'
            crash_items.append(['\u2610', 'REMOVE', your_file,
                                f'OBSOLETE per Mod Checker → {mod_info}',
                                'HIGH', detail])

        if crash_items:
            ws2.cell(row=r, column=1, value="PRIORITY 1 \u2014 FIX CRASHES & BROKEN MODS").font = Font(
                name='Arial', bold=True, size=11, color='D32F2F')
            r += 1
            _header_row(ws2, r, hdrs)
            r += 1
            for vals in crash_items:
                _data_row(ws2, r, vals, RED_FILL)
                r += 1
            r += 1

        # --- P2: SCRIPT MODS ---
        script_items = []
        for item in rows_by_type.get('version_unknown', []):
            notes = item.get('Notes', '')
            mod_name = notes.split(' \u2014 ')[0] if ' \u2014 ' in notes else _basename(item.get('File', ''))
            min_ver = notes.split('\u2265 ')[-1].split(',')[0].strip() if '\u2265' in notes else ''
            script_items.append(['\u2610', 'VERIFY', mod_name,
                                 f'Version undetectable; needs \u2265 {min_ver}' if min_ver else 'Version undetectable',
                                 'LOW', 'Check creator site for latest'])

        if script_items:
            ws2.cell(row=r, column=1, value="PRIORITY 2 \u2014 VERIFY SCRIPT MODS").font = Font(
                name='Arial', bold=True, size=11, color='E65100')
            r += 1
            _header_row(ws2, r, hdrs)
            r += 1
            for vals in script_items:
                _data_row(ws2, r, vals, ORANGE_FILL)
                r += 1
            r += 1

        # --- P3: ANCIENT CC ---
        ancient = rows_by_type.get('outdated_critical', [])
        if ancient:
            ws2.cell(row=r, column=1, value=f"PRIORITY 3 \u2014 REMOVE ANCIENT CC ({len(ancient)} files)").font = Font(
                name='Arial', bold=True, size=11, color='F57F17')
            r += 1
            ws2.cell(row=r, column=1,
                     value="10+ years old. Mostly build/buy CC. Won't crash but may cause visual glitches.").font = SMALL_FONT
            r += 1
            _header_row(ws2, r, ['\u2713', 'File Name', 'Subfolder', 'Date', 'Size (KB)', 'Age'])
            r += 1
            for item in ancient:
                fp = item.get('File', '')
                notes = item.get('Notes', '')
                age = notes.split('(')[-1].rstrip(')') if '(' in notes else ''
                _data_row(ws2, r, ['\u2610', _basename(fp), _subfolder(fp),
                                   item.get('Date', ''),
                                   round(int(item.get('Size', 0)) / 1024),
                                   age], YELLOW_FILL)
                r += 1
            r += 1

        # --- P3B: OUTDATED MEDIUM ---
        outdated = rows_by_type.get('outdated', [])
        if outdated:
            ws2.cell(row=r, column=1, value=f"PRIORITY 3B \u2014 OUTDATED CC ({len(outdated)} files)").font = Font(
                name='Arial', bold=True, size=11, color='F57F17')
            r += 1
            _header_row(ws2, r, ['\u2713', 'File Name', 'Subfolder', 'Date', 'Size (KB)', 'Notes'])
            r += 1
            for item in outdated:
                fp = item.get('File', '')
                _data_row(ws2, r, ['\u2610', _basename(fp), _subfolder(fp),
                                   item.get('Date', ''),
                                   round(int(item.get('Size', 0)) / 1024),
                                   item.get('Notes', '')[:60]], YELLOW_FILL)
                r += 1
            r += 1

        # --- P4: BIG FILES ---
        big = sorted(rows_by_type.get('merged_analysis', []),
                     key=lambda x: int(x.get('Size', 0)), reverse=True)
        if big:
            ws2.cell(row=r, column=1, value="PRIORITY 4 \u2014 OVERSIZED PACKAGES (100MB+)").font = Font(
                name='Arial', bold=True, size=11, color='1565C0')
            r += 1
            ws2.cell(row=r, column=1,
                     value="These eat RAM. Consider removing or finding optimized versions.").font = SMALL_FONT
            r += 1
            _header_row(ws2, r, ['\u2713', 'File Name', 'Subfolder', 'Size (MB)', 'Type', 'Notes'])
            r += 1
            for item in big[:30]:
                fp = item.get('File', '')
                notes = item.get('Notes', '')
                ptype = 'Texture' if 'Texture' in notes else 'Mesh' if 'Mesh' in notes else 'Mixed'
                short = notes.split('|')[-1].strip() if '|' in notes else ''
                _data_row(ws2, r, ['\u2610', _basename(fp), _subfolder(fp),
                                   round(int(item.get('Size', 0)) / (1024 * 1024)),
                                   ptype, short], BLUE_FILL)
                r += 1

        _set_widths(ws2, [4, 16, 55, 42, 8, 55])
        ws2.freeze_panes = 'A4'

        # ── SHEET 3: TUNING CONFLICTS ──
        conflicts = rows_by_type.get('tuning_conflict', [])
        if conflicts:
            ws3 = wb.create_sheet("Tuning Conflicts")
            ws3.sheet_properties.tabColor = "FF9800"
            ws3['A1'] = "TUNING CONFLICTS"
            ws3['A1'].font = TITLE_FONT
            ws3.row_dimensions[1].height = 30
            ws3['A2'] = "Packages overriding the same SimData instance. Only one wins; others are silently ignored."
            ws3['A2'].font = SMALL_FONT
            r3 = 4
            _header_row(ws3, r3, ['File Name', 'Subfolder', 'Date', 'Conflict Instance', 'Risk'])
            r3 += 1
            for i, item in enumerate(conflicts):
                fp = item.get('File', '')
                notes = item.get('Notes', '')
                cid = notes.split('instance ')[-1].split(' ')[0] if 'instance' in notes else ''
                _data_row(ws3, r3, [_basename(fp), _subfolder(fp), item.get('Date', ''), cid,
                                    item.get('Risk', '')], ALT_ROW if i % 2 == 0 else None)
                r3 += 1
            _set_widths(ws3, [55, 22, 14, 24, 10])
            ws3.freeze_panes = 'A5'

        # ── SHEET 4: MOD CHECKER LIST CROSS-REFERENCE ──
        if self.modlist_matches:
            PURPLE_FILL = PatternFill('solid', fgColor='E1BEE7')
            PINK_FILL = PatternFill('solid', fgColor='F8BBD0')
            LIGHT_GREEN = PatternFill('solid', fgColor='C8E6C9')
            LIGHT_BLUE = PatternFill('solid', fgColor='B3E5FC')
            GREY_FILL = PatternFill('solid', fgColor='E0E0E0')

            STATUS_FILLS = {
                'Broken': RED_FILL,
                'Obsolete': PINK_FILL,
                'Minor Issue(s)': ORANGE_FILL,
                'Unknown': YELLOW_FILL,
                'Updated': LIGHT_GREEN,
                'No Known Issues': GREEN_FILL,
                'N/A': GREY_FILL,
                'New': LIGHT_BLUE,
            }

            ws4 = wb.create_sheet("Mod Checker Status")
            ws4.sheet_properties.tabColor = "9C27B0"
            ws4['A1'] = "SCARLET MOD CHECKER — CROSS-REFERENCE"
            ws4['A1'].font = TITLE_FONT
            ws4.row_dimensions[1].height = 30
            ws4['A2'] = (f"Your installed mods matched against the Scarlet Realm Mod Checker list. "
                         f"Generated: {today} | {len(self.modlist_matches)} matches found")
            ws4['A2'].font = SMALL_FONT

            # Summary stats at top
            r4 = 4
            status_counts = defaultdict(int)
            for m in self.modlist_matches:
                status_counts[m['patch_status']] += 1

            _header_row(ws4, r4, ['Status', 'Count', 'Action Needed'])
            r4 += 1
            status_actions = {
                'Broken': 'Remove or update immediately',
                'Obsolete': 'Remove — no longer maintained',
                'Minor Issue(s)': 'Check notes, may need workaround',
                'Unknown': 'Verify manually on creator page',
                'Updated': 'You may need to re-download latest',
                'No Known Issues': 'No action needed',
                'N/A': 'Tracking unavailable',
                'New': 'No action needed',
            }
            for status in ['Broken', 'Obsolete', 'Minor Issue(s)', 'Unknown',
                           'Updated', 'No Known Issues', 'N/A', 'New', '']:
                count = status_counts.get(status, 0)
                if count == 0:
                    continue
                label = status if status else 'No Status'
                fill = STATUS_FILLS.get(status, None)
                action = status_actions.get(status, 'Check manually')
                _data_row(ws4, r4, [label, count, action], fill)
                r4 += 1

            r4 += 1  # Just one blank row before detail table

            # Full detail table — sorted: Broken first, then Obsolete, Minor Issue, Unknown, etc.
            status_order = {'Broken': 0, 'Obsolete': 1, 'Minor Issue(s)': 2,
                            'Unknown': 3, '': 4, 'N/A': 5, 'Updated': 6,
                            'No Known Issues': 7, 'New': 8}
            sorted_matches = sorted(self.modlist_matches,
                                    key=lambda m: (status_order.get(m['patch_status'], 9),
                                                   m['matched_mod'].lower()))

            detail_hdrs = ['\u2713', 'Your File', 'Matched Mod Name', 'Creator',
                           'Patch Status', 'Last Update', 'Status Date', 'Notes']
            _header_row(ws4, r4, detail_hdrs)
            r4 += 1

            for m in sorted_matches:
                fill = STATUS_FILLS.get(m['patch_status'], None)
                # m['file'] is a Path object, m['filename'] is the stem (no extension)
                your_file = m['filename']
                if hasattr(m['file'], 'suffix'):
                    your_file = m['filename'] + m['file'].suffix
                notes_short = m['notes'][:100] if m['notes'] else ''
                _data_row(ws4, r4, [
                    '\u2610', your_file, m['matched_mod'], m['creator'],
                    m['patch_status'], m['last_update'], m['status_date'],
                    notes_short
                ], fill)
                # Don't let rows wrap — set fixed height
                ws4.row_dimensions[r4].height = 18
                r4 += 1

            _set_widths(ws4, [4, 65, 35, 22, 16, 16, 14, 60])
            # Freeze at the header row of the detail table
            detail_header_row = r4 - len(sorted_matches) - 1
            ws4.freeze_panes = f'A{detail_header_row + 1}'

        # Save
        xlsx_path = csv_path.rsplit('.', 1)[0] + '_cleanup_checklist.xlsx'
        try:
            wb.save(xlsx_path)
            print(f"  Exported cleanup checklist to {xlsx_path}")
        except Exception as e:
            print(f"  [Error] Could not save Excel checklist: {e}")

    # ── 13. Save / Load Database ──

    def save_broken_db(self):
        """Save the embedded broken mod database to the Mods folder for user editing."""
        db_path = self.mods_path / "broken_cc_hashes.json"
        data = {
            "known_broken": KNOWN_BROKEN_MODS,
            "notes": (f"Generated by Mod Detector v{VERSION} on "
                      f"{datetime.now().strftime('%Y-%m-%d')}. "
                      f"Add your own entries to extend detection."),
        }
        try:
            # Don't overwrite if it exists and has user additions
            if db_path.exists():
                with open(db_path, "r") as f:
                    existing = json.load(f)
                existing_names = {m["mod_name"] for m in existing.get("known_broken", [])}
                new_entries = [m for m in KNOWN_BROKEN_MODS
                               if m["mod_name"] not in existing_names]
                if new_entries:
                    existing["known_broken"].extend(new_entries)
                    existing["notes"] = (f"Updated by v{VERSION} on "
                                         f"{datetime.now().strftime('%Y-%m-%d')} "
                                         f"(+{len(new_entries)} entries)")
                    with open(db_path, "w") as f:
                        json.dump(existing, f, indent=2)
                    print(f"  Updated {db_path.name} (+{len(new_entries)} new entries)")
                else:
                    print(f"  {db_path.name} is up to date")
            else:
                with open(db_path, "w") as f:
                    json.dump(data, f, indent=2)
                print(f"  Created {db_path.name} ({len(KNOWN_BROKEN_MODS)} entries)")
        except Exception as e:
            print(f"  Warning: Could not save database: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# BATCH FILE GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════

def generate_batch_files(mods_path: str):
    """Create .bat launcher files for common scan modes."""
    script_path = Path(__file__).resolve()
    script_name = script_path.name
    script_dir = script_path.parent
    mods_abs = Path(mods_path).resolve()

    batches = {
        "FULL_SCAN.bat": (
            f"Full Comprehensive Scan v{VERSION}",
            f'python "{script_name}" --mods "{mods_abs}" --full --parallel '
            f'--export "full_scan_report.csv"',
            None
        ),
        "FULL_SCAN_WITH_LOG.bat": (
            f"Full Scan + Exception Log v{VERSION}",
            f'python "{script_name}" --mods "{mods_abs}" --full --parallel '
            f'--log "%LOG_FILE%" --export "full_log_report.csv"',
            (
                'set /p LOG_FILE="Drag & drop your Better Exceptions log here: "\n'
                'if "%LOG_FILE%"=="" (\n'
                '    echo No log file provided, running without log.\n'
                f'    python "{script_name}" --mods "{mods_abs}" --full --parallel '
                f'--export "full_scan_report.csv"\n'
                '    goto :done\n'
                ')\n'
            )
        ),
        "FULL_SCAN_WITH_MODLIST.bat": (
            f"Full Scan + Scarlet Mod Checker v{VERSION}",
            None,  # Custom flow below
            (
                'set /p MODLIST_FILE="Drag & drop your Scarlet Mod Checker CSV here: "\n'
                'if "%MODLIST_FILE%"=="" (\n'
                '    echo No mod list provided, running without mod checker.\n'
                f'    python "{script_name}" --mods "{mods_abs}" --full --parallel '
                f'--export "full_scan_report.csv"\n'
                '    goto :done\n'
                ')\n'
                f'python "{script_name}" --mods "{mods_abs}" --full --parallel '
                f'--modlist "%MODLIST_FILE%" --export "full_modlist_report.csv"\n'
            )
        ),
        "FULL_SCAN_WITH_ALL.bat": (
            f"Full Scan + Log + Mod Checker v{VERSION}",
            None,  # Custom flow below
            (
                'set /p LOG_FILE="Drag & drop your Better Exceptions log here (or press Enter to skip): "\n'
                'set /p MODLIST_FILE="Drag & drop your Scarlet Mod Checker CSV here (or press Enter to skip): "\n'
                'set EXTRA_ARGS=\n'
                'if not "%LOG_FILE%"=="" set EXTRA_ARGS=%EXTRA_ARGS% --log "%LOG_FILE%"\n'
                'if not "%MODLIST_FILE%"=="" set EXTRA_ARGS=%EXTRA_ARGS% --modlist "%MODLIST_FILE%"\n'
                f'python "{script_name}" --mods "{mods_abs}" --full --parallel '
                f'%EXTRA_ARGS% --export "full_complete_report.csv"\n'
            )
        ),
        "QUICK_SCAN.bat": (
            f"Quick Scan v{VERSION}",
            f'python "{script_name}" --mods "{mods_abs}" --fast --parallel '
            f'--export "quick_scan_report.csv"',
            None
        ),
        "MERGED_ANALYSIS.bat": (
            f"Merged Package Analysis v{VERSION}",
            f'python "{script_name}" --mods "{mods_abs}" --analyze-merged '
            f'--fast --parallel --export "merged_analysis_report.csv"',
            None
        ),
        "GUI_LAUNCH.bat": (
            f"GUI Launcher v{VERSION}",
            f'python "{script_name}" --gui',
            None
        ),
        "UPDATE_DATABASE.bat": (
            f"Update Broken Mod Database v{VERSION}",
            f'python "{script_name}" --mods "{mods_abs}" --update-db',
            None
        ),
    }

    print("\n" + "=" * 70)
    print("GENERATING BATCH FILES")
    print("=" * 70)

    for filename, (title, cmd, custom_flow) in batches.items():
        if custom_flow:
            # Custom batch with drag-and-drop prompts
            content = (
                f"@echo off\n"
                f"chcp 65001 >nul\n"
                f"title Sims 4 Mod Detector - {title}\n"
                f"echo {'=' * 60}\n"
                f"echo   {title}\n"
                f"echo {'=' * 60}\n"
                f"echo.\n"
                f'cd /d "{script_dir}"\n'
                f"{custom_flow}"
                f":done\n"
                f"echo.\n"
                f"echo Done!\n"
                f"pause\n"
            )
        else:
            content = (
                f"@echo off\n"
                f"chcp 65001 >nul\n"
                f"title Sims 4 Mod Detector - {title}\n"
                f"echo {'=' * 60}\n"
                f"echo   {title}\n"
                f"echo {'=' * 60}\n"
                f"echo.\n"
                f"echo Press any key to start...\n"
                f"pause >nul\n"
                f'cd /d "{script_dir}"\n'
                f"{cmd}\n"
                f"echo.\n"
                f"echo Done!\n"
                f"pause\n"
            )
        filepath = script_dir / filename
        try:
            with open(filepath, "w", encoding="utf-8") as f:
                f.write(content)
            print(f"  Created: {filename}")
        except Exception as e:
            print(f"  Failed: {filename} — {e}")

    print(f"\n  Batch files saved to: {script_dir}")


# ═══════════════════════════════════════════════════════════════════════════════
# GUI (Tkinter)
# ═══════════════════════════════════════════════════════════════════════════════

def launch_gui():
    """Launch the Tkinter GUI."""
    try:
        import tkinter as tk
        from tkinter import filedialog, scrolledtext, messagebox
        from tkinter.ttk import Button, Label, Checkbutton
    except ImportError:
        print("Error: tkinter not available. Use CLI mode instead.")
        sys.exit(1)

    class ModDetectorGUI:
        def __init__(self, root):
            self.root = root
            self.root.title(f"Sims 4 Mod Detector v{VERSION}")
            self.root.geometry("950x750")
            self.analyzer = None
            self._build_ui()

        def _build_ui(self):
            # ── Mods folder ──
            Label(self.root, text="Mods Folder:").grid(
                row=0, column=0, sticky="w", padx=10, pady=5)
            self.mods_var = tk.StringVar(
                value=str(Path.home() / "Documents" / "Electronic Arts"
                          / "The Sims 4" / "Mods"))
            tk.Entry(self.root, textvariable=self.mods_var, width=65).grid(
                row=0, column=1, padx=5, pady=5)
            Button(self.root, text="Browse",
                   command=self._browse_mods).grid(row=0, column=2, padx=5)

            # ── Log file ──
            Label(self.root, text="Log File (optional):").grid(
                row=1, column=0, sticky="w", padx=10, pady=5)
            self.log_var = tk.StringVar()
            tk.Entry(self.root, textvariable=self.log_var, width=65).grid(
                row=1, column=1, padx=5, pady=5)
            Button(self.root, text="Browse",
                   command=self._browse_log).grid(row=1, column=2, padx=5)

            # ── Mod Checker List ──
            Label(self.root, text="Mod Checker CSV (optional):").grid(
                row=2, column=0, sticky="w", padx=10, pady=5)
            self.modlist_var = tk.StringVar()
            tk.Entry(self.root, textvariable=self.modlist_var, width=65).grid(
                row=2, column=1, padx=5, pady=5)
            Button(self.root, text="Browse",
                   command=self._browse_modlist).grid(row=2, column=2, padx=5)

            # ── Options ──
            self.fast_var = tk.BooleanVar(value=False)
            self.parallel_var = tk.BooleanVar(value=True)
            tk.Checkbutton(self.root, text="Fast Mode (skip heavy checks)",
                           variable=self.fast_var).grid(row=3, column=0, sticky="w", padx=10)
            tk.Checkbutton(self.root, text="Parallel Hashing",
                           variable=self.parallel_var).grid(row=3, column=1, sticky="w")

            # ── Buttons ──
            Button(self.root, text="Run Scan",
                   command=self._run_scan).grid(row=4, column=0, pady=10, ipadx=20, ipady=8)
            Button(self.root, text="Export CSV",
                   command=self._export_csv).grid(row=4, column=1, pady=10, ipadx=20, ipady=8)
            Button(self.root, text="Generate .bat Files",
                   command=self._gen_batch).grid(row=4, column=2, pady=10, ipadx=10, ipady=8)

            # ── Results ──
            Label(self.root, text=f"Results (v{VERSION}):").grid(
                row=5, column=0, columnspan=3, sticky="w", padx=10, pady=5)
            self.results_text = scrolledtext.ScrolledText(
                self.root, width=110, height=28, wrap=tk.WORD, font=("Consolas", 9))
            self.results_text.grid(row=6, column=0, columnspan=3, padx=10, pady=5)

        def _browse_mods(self):
            folder = filedialog.askdirectory(title="Select Mods Folder")
            if folder:
                self.mods_var.set(folder)

        def _browse_log(self):
            f = filedialog.askopenfilename(
                title="Select Exception Log",
                filetypes=[("Text/HTML", "*.txt *.html"), ("All", "*.*")])
            if f:
                self.log_var.set(f)

        def _browse_modlist(self):
            f = filedialog.askopenfilename(
                title="Select Scarlet Mod Checker CSV",
                filetypes=[("CSV", "*.csv"), ("All", "*.*")])
            if f:
                self.modlist_var.set(f)

        def _capture_run(self, func):
            """Run a function, capturing stdout to the results pane."""
            from io import StringIO
            old = sys.stdout
            sys.stdout = capture = StringIO()
            try:
                func()
            except Exception as e:
                print(f"\nERROR: {e}")
            finally:
                sys.stdout = old
            self.results_text.insert(tk.END, capture.getvalue())
            self.results_text.see(tk.END)
            self.root.update()

        def _run_scan(self):
            self.results_text.delete("1.0", tk.END)
            mods = self.mods_var.get()
            log = self.log_var.get() or None
            modlist = self.modlist_var.get() or None
            fast = self.fast_var.get()
            parallel = self.parallel_var.get()

            if not mods:
                messagebox.showerror("Error", "Select a Mods folder first.")
                return

            self.analyzer = ModAnalyzer(mods, log, modlist)

            def do_scan():
                if not self.analyzer.scan_mods():
                    print("ERROR: Mods folder not found.")
                    return
                # Load mod checker list if provided
                if modlist:
                    self.analyzer.load_modlist()
                self.analyzer.check_integrity()
                self.analyzer.detect_duplicates(parallel=parallel)
                self.analyzer.detect_outdated()
                self.analyzer.detect_broken_mods()
                # Cross-reference with mod checker list
                if self.analyzer.modlist_db:
                    self.analyzer.cross_reference_modlist()
                if not fast:
                    self.analyzer.detect_tuning_conflicts()
                    self.analyzer.analyze_mesh_quality()
                self.analyzer.profile_performance()
                if not fast:
                    self.analyzer.detect_mod_versions()
                    self.analyzer.analyze_exception_log()
                    self.analyzer.analyze_resource_composition()
                    self.analyzer.analyze_merged_packages(threshold_mb=100.0)
                self.analyzer.generate_report()
                print(f"\nScan complete. {len(self.analyzer.issues)} issues found.")

            self._capture_run(do_scan)

        def _export_csv(self):
            if not self.analyzer or not self.analyzer.issues:
                messagebox.showerror("Error", "Run a scan first.")
                return
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV", "*.csv"), ("All", "*.*")])
            if filename:
                self.analyzer.export_to_csv(filename)
                messagebox.showinfo("Success", f"Exported to {filename}")

        def _gen_batch(self):
            mods = self.mods_var.get()
            if not mods:
                messagebox.showerror("Error", "Select a Mods folder first.")
                return
            self.results_text.delete("1.0", tk.END)
            self._capture_run(lambda: generate_batch_files(mods))

    root = tk.Tk()
    ModDetectorGUI(root)
    root.mainloop()


# ═══════════════════════════════════════════════════════════════════════════════
# CLI ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description=f"Sims 4 Mod Conflict Detector v{VERSION}",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=f"""
Version {VERSION} | {BUILD_DATE}

Capabilities:
  - Proper DBPF 2.0/2.1 package parsing (real TGI index reading)
  - Tuning conflict detection via resource instance IDs
  - LOD quality analysis from actual GEOM resource counts
  - DBPF header integrity validation
  - Known broken mod database (embedded + local JSON extensible)
  - Scarlet Mod Checker list cross-reference (broken/updated/obsolete)
  - Script mod version detection (reads inside .ts4script and .package)
  - Merged package deep analysis (resource composition breakdown)
  - Duplicate detection (hash + filename similarity)
  - Exception log analysis (Better Exceptions / MCCC)
  - Performance profiling (file sizes, resource counts, script load)
  - CSV export with risk levels

Examples:
  Full scan:        python {Path(__file__).name} --mods "path/to/Mods" --full
  With mod list:    python {Path(__file__).name} --mods "path/to/Mods" --modlist "Mod_List_Checker.csv" --full
  Quick scan:       python {Path(__file__).name} --mods "path/to/Mods" --fast --parallel
  With log:         python {Path(__file__).name} --mods "path/to/Mods" --log "path/to/log.txt"
  Merged analysis:  python {Path(__file__).name} --mods "path/to/Mods" --analyze-merged
  GUI mode:         python {Path(__file__).name} --gui
  Make .bat files:  python {Path(__file__).name} --mods "path/to/Mods" --generate-batch
        """,
    )

    parser.add_argument("--mods", "-m", help="Path to Sims 4 Mods folder")
    parser.add_argument("--log", "-l", help="Path to exception log file (optional)")
    parser.add_argument("--modlist", help="Path to Scarlet Mod Checker CSV (optional)")
    parser.add_argument("--export", help="Export CSV to this file (default: mod_report.csv)")
    parser.add_argument("--gui", action="store_true", help="Launch GUI")
    parser.add_argument("--fast", action="store_true",
                        help="Fast mode: skip DBPF parsing, tuning conflicts, LOD analysis")
    parser.add_argument("--full", action="store_true",
                        help="Full scan: all checks enabled, parallel on")
    parser.add_argument("--parallel", action="store_true", help="Parallel file hashing")
    parser.add_argument("--update-db", action="store_true",
                        help="Save/update broken mod database JSON in Mods folder")
    parser.add_argument("--generate-batch", action="store_true",
                        help="Generate .bat launcher files")
    parser.add_argument("--analyze-merged", action="store_true",
                        help="Deep analysis of merged packages (>100 MB)")
    parser.add_argument("--merged-threshold", type=float, default=100.0,
                        help="Size threshold in MB for merged analysis (default: 100)")

    args = parser.parse_args()

    if args.gui:
        launch_gui()
        return

    if not args.mods:
        print(f"Error: --mods is required for CLI mode (or use --gui)")
        parser.print_help()
        sys.exit(1)

    print("=" * 70)
    print(f"  Sims 4 Mod Detector v{VERSION}")
    print(f"  DBPF parsing | TGI conflicts | LOD analysis | Broken mod DB")
    print(f"  Build: {BUILD_DATE}")
    print("=" * 70)

    if args.generate_batch:
        generate_batch_files(args.mods)
        return

    if args.full:
        args.parallel = True
        args.fast = False

    analyzer = ModAnalyzer(args.mods, args.log, args.modlist)

    if args.update_db:
        print("\nUpdating broken mod database...")
        analyzer.save_broken_db()

    if not analyzer.scan_mods():
        sys.exit(1)

    # ── Load mod checker list if provided ──
    if args.modlist:
        analyzer.load_modlist()

    # ── Run checks ──
    analyzer.check_integrity()
    analyzer.detect_duplicates(parallel=args.parallel)
    analyzer.detect_outdated()
    analyzer.detect_broken_mods()

    # ── Cross-reference with mod checker list ──
    if analyzer.modlist_db:
        analyzer.cross_reference_modlist()

    if not args.fast:
        analyzer.detect_tuning_conflicts()
        analyzer.analyze_mesh_quality()

    analyzer.profile_performance()

    if not args.fast:
        analyzer.detect_mod_versions()
        analyzer.analyze_exception_log()
        analyzer.analyze_resource_composition()

    # Merged analysis: always run on --full, or on explicit flag
    if args.analyze_merged or args.full:
        analyzer.analyze_merged_packages(threshold_mb=args.merged_threshold)

    analyzer.generate_report()

    # ── Export ──
    export_path = args.export or "mod_report.csv"
    print()
    analyzer.export_to_csv(export_path)
    analyzer.export_cleanup_xlsx(export_path)

    print("\n" + "=" * 70)
    print(f"  SCAN COMPLETE")
    print("=" * 70)
    print(f"  CSV Report:       {export_path}")
    xlsx_path = export_path.rsplit('.', 1)[0] + '_cleanup_checklist.xlsx'
    if os.path.exists(xlsx_path):
        print(f"  Excel Checklist:  {xlsx_path}")
    print(f"  Backup your Mods folder before making changes!")
    if not args.generate_batch:
        print(f"  Tip: Run with --generate-batch to create .bat launchers")


if __name__ == "__main__":
    freeze_support()
    main()
